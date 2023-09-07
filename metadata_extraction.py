#*- coding: utf-8 -*-
import csv
import re
import os
import time
import pymarc
from bs4 import BeautifulSoup
import urllib.request
import tkinter
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import asksaveasfilename
from tkinter.filedialog import askdirectory
from tkinter import ttk
from tkinter import font
from functools import partial
import configparser
from collections import OrderedDict
import openpyxl
import win32clipboard
import pywintypes
import tifffile
from utilities.crl_folio_utilities import *
import utilities.text_marc_reader as text_marc_reader

#Creates the folder if it does not already exist
def check_or_create_dir(path):
    if not os.path.exists(path):
        if not os.path.exists(os.path.dirname(path)):
            check_or_create_dir(os.path.dirname(path))
        os.mkdir(path)

file_direct = os.path.dirname(os.path.realpath('__file__'))
output_folder = os.path.join(file_direct, 'Output')
check_or_create_dir(output_folder)

country_dict = {'sdu' : 'United States', 'xs' : 'South Georgia and the South Sandwich Islands', 'sd' : 'South Sudan', 'sp' : 'Spain', 'sh' : 'Spanish North Africa', 'xp' : 'Spratly Island', 'ce' : 'Sri Lanka', 'sj' : 'Sudan', 'sr' : 'Surinam', 'sq' : 'Swaziland', 'sw' : 'Sweden', 'sz' : 'Switzerland', 'sy' : 'Syria', 'ta' : 'Tajikistan', 'tz' : 'Tanzania', 'tma' : 'Australia', 'tnu' : 'United States', 'fs' : 'Terres australes et antarctiques françaises', 'txu' : 'United States', 'th' : 'Thailand', 'em' : 'Timor-Leste', 'tg' : 'Togo', 'tl' : 'Tokelau', 'to' : 'Tonga', 'tr' : 'Trinidad and Tobago', 'ti' : 'Tunisia', 'tu' : 'Turkey', 'tk' : 'Turkmenistan', 'tc' : 'Turks and Caicos Islands', 'tv' : 'Tuvalu', 'ug' : 'Uganda', 'un' : 'Ukraine', 'ts' : 'United Arab Emirates', 'xxk' : 'United Kingdom', 'xxu' : 'United States', 'uc' : 'United States Misc. Caribbean Islands', 'up' : 'United States Misc. Pacific Islands', 'uy' : 'Uruguay', 'utu' : 'United States', 'uz' : 'Uzbekistan', 'nn' : 'Vanuatu', 'vp' : 'Various places', 'vc' : 'Vatican City', 've' : 'Venezuela', 'vtu' : 'United States', 'vra' : 'Australia', 'vm' : 'Vietnam', 'vi' : 'Virgin Islands of the United States', 'vau' : 'United States', 'wk' : 'Wake Island', 'wlk' : 'United Kingdom', 'wf' : 'Wallis and Futuna', 'wau' : 'United States', 'wj' : 'West Bank of the Jordan River', 'wvu' : 'United States', 'wea' : 'Australia', 'ss' : 'Western Sahara', 'wiu' : 'United States', 'wyu' : 'United States', 'ye' : 'Yemen', 'ykc' : 'Canada', 'za' : 'Zambia', 'rh' : 'Zimbabwe', 'ac' : 'Australia', 'ai' : 'Armenia (Republic)', 'air' : 'Armenia (Republic)', 'ajr' : 'Azerbaijan', 'bwr' : 'Belarus', 'cn' : 'Canada', 'cp' : 'Kiribati', 'cs' : 'Czechoslovakia', 'cz' : 'Panama', 'err' : 'Estonia', 'ge' : 'Germany', 'gn' : 'Kiribati', 'gsr' : 'Georgia (Republic)', 'hk' : 'China', 'iu' : 'Israel', 'iw' : 'Israel', 'jn' : 'Norway', 'kgr' : 'Kyrgyzstan', 'kzr' : 'Kazakhstan', 'lir' : 'Lithuania', 'ln' : 'Kiribati', 'lvr' : 'Latvia', 'mh' : 'China', 'mvr' : 'Moldova', 'na' : 'Netherlands Antilles', 'nm' : 'Northern Mariana Islands', 'pt' : 'Timor-Leste', 'rur' : 'Russia (Federation)', 'ry' : 'Japan', 'sb' : 'Norway', 'sk' : 'India', 'sv' : 'Honduras', 'tar' : 'Tajikistan', 'tkr' : 'Turkmenistan', 'tt' : 'Trust Territory of the Pacific Islands', 'ui' : 'United Kingdom Misc. Islands', 'uik' : 'United Kingdom Misc. Islands', 'uk' : 'United Kingdom', 'unr' : 'Ukraine', 'ur' : 'Soviet Union', 'us' : 'United States', 'uzr' : 'Uzbekistan', 'vn' : 'Vietnam', 'vs' : 'Vietnam', 'wb' : 'Germany', 'xi' : 'Saint Kitts-Nevis-Anguilla', 'xxr' : 'Soviet Union', 'ys' : 'Yemen', 'yu' : 'Serbia and Montenegro', 'aa' : 'Albania', 'abc' : 'Canada', 'aca' : 'Australia', 'ae' : 'Algeria', 'af' : 'Afghanistan', 'ag' : 'Argentina', 'aj' : 'Azerbaijan', 'aku' : 'United States', 'alu' : 'United States', 'am' : 'Anguilla', 'an' : 'Andorra', 'ao' : 'Angola', 'aq' : 'Antigua and Barbuda', 'aru' : 'United States', 'as' : 'American Samoa', 'at' : 'Australia', 'au' : 'Austria', 'aw' : 'Aruba', 'ay' : 'Antarctica', 'azu' : 'United States', 'ba' : 'Bahrain', 'bb' : 'Barbados', 'bcc' : 'Canada', 'bd' : 'Burundi', 'be' : 'Belgium', 'bf' : 'Bahamas', 'bg' : 'Bangladesh', 'bh' : 'Belize', 'bi' : 'British Indian Ocean Territory', 'bl' : 'Brazil', 'bm' : 'Bermuda Islands', 'bn' : 'Bosnia and Herzegovina', 'bo' : 'Bolivia', 'bp' : 'Solomon Islands', 'br' : 'Burma', 'bs' : 'Botswana', 'bt' : 'Bhutan', 'bu' : 'Bulgaria', 'bv' : 'Bouvet Island', 'bw' : 'Belarus', 'bx' : 'Brunei', 'ca' : 'Caribbean Netherlands', 'cau' : 'United States', 'cb' : 'Cambodia', 'cc' : 'China', 'cd' : 'Chad', 'cf' : 'Congo (Brazzaville)', 'cg' : 'Congo (Democratic Republic)', 'ch' : 'China (Republic : 1949- )', 'ci' : 'Croatia', 'cj' : 'Cayman Islands', 'ck' : 'Colombia', 'cl' : 'Chile', 'cm' : 'Cameroon', 'co' : 'Curaçao', 'cou' : 'United States', 'cq' : 'Comoros', 'cr' : 'Costa Rica', 'ctu' : 'United States', 'cu' : 'Cuba', 'cv' : 'Cabo Verde', 'cw' : 'Cook Islands', 'cx' : 'Central African Republic', 'cy' : 'Cyprus', 'dcu' : 'United States', 'deu' : 'United States', 'dk' : 'Denmark', 'dm' : 'Benin', 'dq' : 'Dominica', 'dr' : 'Dominican Republic', 'ea' : 'Eritrea', 'ec' : 'Ecuador', 'eg' : 'Equatorial Guinea', 'enk' : 'United Kingdom', 'er' : 'Estonia', 'es' : 'El Salvador', 'et' : 'Ethiopia', 'fa' : 'Faroe Islands', 'fg' : 'French Guiana', 'fi' : 'Finland', 'fj' : 'Fiji', 'fk' : 'Falkland Islands', 'flu' : 'United States', 'fm' : 'Micronesia (Federated States)', 'fp' : 'French Polynesia', 'fr' : 'France', 'ft' : 'Djibouti', 'gau' : 'United States', 'gb' : 'Kiribati', 'gd' : 'Grenada', 'gg' : 'Guernsey', 'gh' : 'Ghana', 'gi' : 'Gibraltar', 'gl' : 'Greenland', 'gm' : 'Gambia', 'go' : 'Gabon', 'gp' : 'Guadeloupe', 'gr' : 'Greece', 'gs' : 'Georgia (Republic)', 'gt' : 'Guatemala', 'gu' : 'Guam', 'gv' : 'Guinea', 'gw' : 'Germany', 'gy' : 'Guyana', 'gz' : 'Gaza Strip', 'hiu' : 'United States', 'hm' : 'Heard and McDonald Islands', 'ho' : 'Honduras', 'ht' : 'Haiti', 'hu' : 'Hungary', 'iau' : 'United States', 'ic' : 'Iceland', 'idu' : 'United States', 'ie' : 'Ireland', 'ii' : 'India', 'ilu' : 'United States', 'im' : 'Isle of Man', 'inu' : 'United States', 'io' : 'Indonesia', 'iq' : 'Iraq', 'ir' : 'Iran', 'is' : 'Israel', 'it' : 'Italy', 'iv' : 'Côte d\'Ivoire', 'iy' : 'Iraq-Saudi Arabia Neutral Zone', 'ja' : 'Japan', 'je' : 'Jersey', 'ji' : 'Johnston Atoll', 'jm' : 'Jamaica', 'jo' : 'Jordan', 'ke' : 'Kenya', 'kg' : 'Kyrgyzstan', 'kn' : 'Korea (North)', 'ko' : 'Korea (South)', 'ksu' : 'United States', 'ku' : 'Kuwait', 'kv' : 'Kosovo', 'kyu' : 'United States', 'kz' : 'Kazakhstan', 'lau' : 'United States', 'lb' : 'Liberia', 'le' : 'Lebanon', 'lh' : 'Liechtenstein', 'li' : 'Lithuania', 'lo' : 'Lesotho', 'ls' : 'Laos', 'lu' : 'Luxembourg', 'lv' : 'Latvia', 'ly' : 'Libya', 'mau' : 'United States', 'mbc' : 'Canada', 'mc' : 'Monaco', 'mdu' : 'United States', 'meu' : 'United States', 'mf' : 'Mauritius', 'mg' : 'Madagascar', 'miu' : 'United States', 'mj' : 'Montserrat', 'mk' : 'Oman', 'ml' : 'Mali', 'mm' : 'Malta', 'mnu' : 'United States', 'mo' : 'Montenegro', 'mou' : 'United States', 'mp' : 'Mongolia', 'mq' : 'Martinique', 'mr' : 'Morocco', 'msu' : 'United States', 'mtu' : 'United States', 'mu' : 'Mauritania', 'mv' : 'Moldova', 'mw' : 'Malawi', 'mx' : 'Mexico', 'my' : 'Malaysia', 'mz' : 'Mozambique', 'nbu' : 'United States', 'ncu' : 'United States', 'ndu' : 'United States', 'ne' : 'Netherlands', 'nfc' : 'Canada', 'ng' : 'Niger', 'nhu' : 'United States', 'nik' : 'United Kingdom', 'nju' : 'United States', 'nkc' : 'Canada', 'nl' : 'New Caledonia', 'nmu' : 'United States', 'no' : 'Norway', 'np' : 'Nepal', 'nq' : 'Nicaragua', 'nr' : 'Nigeria', 'nsc' : 'Canada', 'ntc' : 'Canada', 'nu' : 'Nauru', 'nuc' : 'Canada', 'nvu' : 'United States', 'nw' : 'Northern Mariana Islands', 'nx' : 'Norfolk Island', 'nyu' : 'United States', 'nz' : 'New Zealand', 'ohu' : 'United States', 'oku' : 'United States', 'onc' : 'Canada', 'oru' : 'United States', 'ot' : 'Mayotte', 'pau' : 'United States', 'pc' : 'Pitcairn Island', 'pe' : 'Peru', 'pf' : 'Paracel Islands', 'pg' : 'Guinea-Bissau', 'ph' : 'Philippines', 'pic' : 'Canada', 'pk' : 'Pakistan', 'pl' : 'Poland', 'pn' : 'Panama', 'po' : 'Portugal', 'pp' : 'Papua New Guinea', 'pr' : 'Puerto Rico', 'pw' : 'Palau', 'py' : 'Paraguay', 'qa' : 'Qatar', 'qea' : 'Queensland', 'quc' : 'Canada', 'rb' : 'Serbia', 're' : 'Réunion', 'riu' : 'United States', 'rm' : 'Romania', 'ru' : 'Russia (Federation)', 'rw' : 'Rwanda', 'sa' : 'South Africa', 'sc' : 'Saint-Barthélemy', 'scu' : 'United States', 'se' : 'Seychelles', 'sf' : 'Sao Tome and Principe', 'sg' : 'Senegal', 'si' : 'Singapore', 'sl' : 'Sierra Leone', 'sm' : 'San Marino', 'sn' : 'Sint Maarten', 'snc' : 'Canada', 'so' : 'Somalia', 'st' : 'Saint-Martin', 'stk' : 'United Kingdom', 'su' : 'Saudi Arabia', 'sx' : 'Namibia', 'tv' : 'Tuvalu', 'ua' : 'Egypt', 'uv' : 'Burkina Faso', 'vb' : 'British Virgin Islands', 'vi' : 'Virgin Islands of the United States', 'ws' : 'Samoa', 'xa' : 'Christmas Island (Indian Ocean)', 'xb' : 'Cocos (Keeling) Islands', 'xc' : 'Maldives', 'xd' : 'Saint Kitts-Nevis', 'xe' : 'Marshall Islands', 'xf' : 'Midway Islands', 'xga' : 'Coral Sea Islands Territory', 'xh' : 'Niue', 'xj' : 'Saint Helena', 'xk' : 'Saint Lucia', 'xl' : 'Saint Pierre and Miquelon', 'xm' : 'Saint Vincent and the Grenadines', 'xn' : 'Macedonia', 'xna' : 'New South Wales', 'xo' : 'Slovakia', 'xoa' : 'Northern Territory', 'xr' : 'Czech Republic', 'xra' : 'South Australia', 'xv' : 'Slovenia', 'xx' : 'No place, unknown, or undetermined', 'xxc' : 'Canada', 'xxu' : 'United States'}
language_dict = {'aar' : 'Afar', 'abk' : 'Abkhaz', 'ace' : 'Achinese', 'ach' : 'Acoli', 'ada' : 'Adangme', 'ady' : 'Adygei', 'afa' : 'Afroasiatic (Other)', 'afh' : 'Afrihili (Artificial language)', 'afr' : 'Afrikaans', 'ain' : 'Ainu', 'ajm' : 'Aljamía', 'aka' : 'Akan', 'akk' : 'Akkadian', 'alb' : 'Albanian', 'ale' : 'Aleut', 'alg' : 'Algonquian (Other)', 'alt' : 'Altai', 'amh' : 'Amharic', 'ang' : 'English, Old (ca. 450-1100)', 'anp' : 'Angika', 'apa' : 'Apache languages', 'ara' : 'Arabic', 'arc' : 'Aramaic', 'arg' : 'Aragonese', 'arm' : 'Armenian', 'arn' : 'Mapuche', 'arp' : 'Arapaho', 'art' : 'Artificial (Other)', 'arw' : 'Arawak', 'asm' : 'Assamese', 'ast' : 'Bable', 'ath' : 'Athapascan (Other)', 'aus' : 'Australian languages', 'ava' : 'Avaric', 'ave' : 'Avestan', 'awa' : 'Awadhi', 'aym' : 'Aymara', 'aze' : 'Azerbaijani', 'bad' : 'Banda languages', 'bai' : 'Bamileke languages', 'bak' : 'Bashkir', 'bal' : 'Baluchi', 'bam' : 'Bambara', 'ban' : 'Balinese', 'baq' : 'Basque', 'bas' : 'Basa', 'bat' : 'Baltic (Other)', 'bej' : 'Beja', 'bel' : 'Belarusian', 'bem' : 'Bemba', 'ben' : 'Bengali', 'ber' : 'Berber (Other)', 'bho' : 'Bhojpuri', 'bih' : 'Bihari (Other) ', 'bik' : 'Bikol', 'bin' : 'Edo', 'bis' : 'Bislama', 'bla' : 'Siksika', 'bnt' : 'Bantu (Other)', 'bos' : 'Bosnian', 'bra' : 'Braj', 'bre' : 'Breton', 'btk' : 'Batak', 'bua' : 'Buriat', 'bug' : 'Bugis', 'bul' : 'Bulgarian', 'bur' : 'Burmese', 'byn' : 'Bilin', 'cad' : 'Caddo', 'cai' : 'Central American Indian (Other)', 'cam' : 'Khmer', 'car' : 'Carib', 'cat' : 'Catalan', 'cau' : 'Caucasian (Other)', 'ceb' : 'Cebuano', 'cel' : 'Celtic (Other)', 'cha' : 'Chamorro', 'chb' : 'Chibcha', 'che' : 'Chechen', 'chg' : 'Chagatai', 'chi' : 'Chinese', 'chk' : 'Chuukese', 'chm' : 'Mari', 'chn' : 'Chinook jargon', 'cho' : 'Choctaw', 'chp' : 'Chipewyan', 'chr' : 'Cherokee', 'chu' : 'Church Slavic', 'chv' : 'Chuvash', 'chy' : 'Cheyenne', 'cmc' : 'Chamic languages', 'cnr' : 'Montenegrin', 'cop' : 'Coptic', 'cor' : 'Cornish', 'cos' : 'Corsican', 'cpe' : 'Creoles and Pidgins, English-based (Other)', 'cpf' : 'Creoles and Pidgins, French-based (Other)', 'cpp' : 'Creoles and Pidgins, Portuguese-based (Other)', 'cre' : 'Cree', 'crh' : 'Crimean Tatar', 'crp' : 'Creoles and Pidgins (Other)', 'csb' : 'Kashubian', 'cus' : 'Cushitic (Other)', 'cze' : 'Czech', 'dak' : 'Dakota', 'dan' : 'Danish', 'dar' : 'Dargwa', 'day' : 'Dayak', 'del' : 'Delaware', 'den' : 'Slavey', 'dgr' : 'Dogrib', 'din' : 'Dinka', 'div' : 'Divehi', 'doi' : 'Dogri', 'dra' : 'Dravidian (Other)', 'dsb' : 'Lower Sorbian', 'dua' : 'Duala', 'dum' : 'Dutch, Middle (ca. 1050-1350)', 'dut' : 'Dutch', 'dyu' : 'Dyula', 'dzo' : 'Dzongkha', 'efi' : 'Efik', 'egy' : 'Egyptian', 'eka' : 'Ekajuk', 'elx' : 'Elamite', 'eng' : 'English', 'enm' : 'English, Middle (1100-1500)', 'epo' : 'Esperanto', 'esk' : 'Eskimo languages', 'esp' : 'Esperanto', 'est' : 'Estonian', 'eth' : 'Ethiopic', 'ewe' : 'Ewe', 'ewo' : 'Ewondo', 'fan' : 'Fang', 'fao' : 'Faroese', 'far' : 'Faroese', 'fat' : 'Fanti', 'fij' : 'Fijian', 'fil' : 'Filipino', 'fin' : 'Finnish', 'fiu' : 'Finno-Ugrian (Other)', 'fon' : 'Fon', 'fre' : 'French', 'fri' : 'Frisian', 'frm' : 'French, Middle (ca. 1300-1600)', 'fro' : 'French, Old (ca. 842-1300)', 'frr' : 'North Frisian', 'frs' : 'East Frisian', 'fry' : 'Frisian', 'ful' : 'Fula', 'fur' : 'Friulian', 'gaa' : 'Gã', 'gae' : 'Scottish Gaelix', 'gag' : 'Galician', 'gal' : 'Oromo', 'gay' : 'Gayo', 'gba' : 'Gbaya', 'gem' : 'Germanic (Other)', 'geo' : 'Georgian', 'ger' : 'German', 'gez' : 'Ethiopic', 'gil' : 'Gilbertese', 'gla' : 'Scottish Gaelic', 'gle' : 'Irish', 'glg' : 'Galician', 'glv' : 'Manx', 'gmh' : 'German, Middle High (ca. 1050-1500)', 'goh' : 'German, Old High (ca. 750-1050)', 'gon' : 'Gondi', 'gor' : 'Gorontalo', 'got' : 'Gothic', 'grb' : 'Grebo', 'grc' : 'Greek, Ancient (to 1453)', 'gre' : 'Greek, Modern (1453-)', 'grn' : 'Guarani', 'gsw' : 'Swiss German', 'gua' : 'Guarani', 'guj' : 'Gujarati', 'gwi' : 'Gwich\'in', 'hai' : 'Haida', 'hat' : 'Haitian French Creole', 'hau' : 'Hausa', 'haw' : 'Hawaiian', 'heb' : 'Hebrew', 'her' : 'Herero', 'hil' : 'Hiligaynon', 'him' : 'Western Pahari languages', 'hin' : 'Hindi', 'hit' : 'Hittite', 'hmn' : 'Hmong', 'hmo' : 'Hiri Motu', 'hrv' : 'Croatian', 'hsb' : 'Upper Sorbian', 'hun' : 'Hungarian', 'hup' : 'Hupa', 'iba' : 'Iban', 'ibo' : 'Igbo', 'ice' : 'Icelandic', 'ido' : 'Ido', 'iii' : 'Sichuan Yi', 'ijo' : 'Ijo', 'iku' : 'Inuktitut', 'ile' : 'Interlingue', 'ilo' : 'Iloko', 'ina' : 'Interlingua (International Auxiliary Language Association)', 'inc' : 'Indic (Other)', 'ind' : 'Indonesian', 'ine' : 'Indo-European (Other)', 'inh' : 'Ingush', 'int' : 'Interlingua (International Auxiliary Language Association)', 'ipk' : 'Inupiaq', 'ira' : 'Iranian (Other)', 'iri' : 'Irish', 'iro' : 'Iroquoian (Other)', 'ita' : 'Italian', 'jav' : 'Javanese', 'jbo' : 'Lojban (Artificial language)', 'jpn' : 'Japanese', 'jpr' : 'Judeo-Persian', 'jrb' : 'Judeo-Arabic', 'kaa' : 'Kara-Kalpak', 'kab' : 'Kabyle', 'kac' : 'Kachin', 'kal' : 'Kalâtdlisut', 'kam' : 'Kamba', 'kan' : 'Kannada', 'kar' : 'Karen languages', 'kas' : 'Kashmiri', 'kau' : 'Kanuri', 'kaw' : 'Kawi', 'kaz' : 'Kazakh', 'kbd' : 'Kabardian', 'kha' : 'Khasi', 'khi' : 'Khoisan (Other)', 'khm' : 'Khmer', 'kho' : 'Khotanese', 'kik' : 'Kikuyu', 'kin' : 'Kinyarwanda', 'kir' : 'Kyrgyz', 'kmb' : 'Kimbundu', 'kok' : 'Konkani', 'kom' : 'Komi', 'kon' : 'Kongo', 'kor' : 'Korean', 'kos' : 'Kosraean', 'kpe' : 'Kpelle', 'krc' : 'Karachay-Balkar', 'krl' : 'Karelian', 'kro' : 'Kru (Other)', 'kru' : 'Kurukh', 'kua' : 'Kuanyama', 'kum' : 'Kumyk', 'kur' : 'Kurdish', 'kus' : 'Kusaie', 'kut' : 'Kootenai', 'lad' : 'Ladino', 'lah' : 'Lahndā', 'lam' : 'Lamba (Zambia and Congo)', 'lan' : 'Occitan (post 1500)', 'lao' : 'Lao', 'lap' : 'Sami', 'lat' : 'Latin', 'lav' : 'Latvian', 'lez' : 'Lezgian', 'lim' : 'Limburgish', 'lin' : 'Lingala', 'lit' : 'Lithuanian', 'lol' : 'Mongo-Nkundu', 'loz' : 'Lozi', 'ltz' : 'Luxembourgish', 'lua' : 'Luba-Lulua', 'lub' : 'Luba-Katanga', 'lug' : 'Ganda', 'lui' : 'Luiseño', 'lun' : 'Lunda', 'luo' : 'Luo (Kenya and Tanzania)', 'lus' : 'Lushai', 'mac' : 'Macedonian', 'mad' : 'Madurese', 'mag' : 'Magahi', 'mah' : 'Marshallese', 'mai' : 'Maithili', 'mak' : 'Makasar', 'mal' : 'Malayalam', 'man' : 'Mandingo', 'mao' : 'Maori', 'map' : 'Austronesian (Other)', 'mar' : 'Marathi', 'mas' : 'Maasai', 'max' : 'Manx', 'may' : 'Malay', 'mdf' : 'Moksha', 'mdr' : 'Mandar', 'men' : 'Mende', 'mga' : 'Irish, Middle (ca. 1100-1550)', 'mic' : 'Micmac', 'min' : 'Minangkabau', 'mis' : 'Miscellaneous languages', 'mkh' : 'Mon-Khmer (Other)', 'mla' : 'Malagasy', 'mlg' : 'Malagasy', 'mlt' : 'Maltese', 'mnc' : 'Manchu', 'mni' : 'Manipuri', 'mno' : 'Manobo languages', 'moh' : 'Mohawk', 'mol' : 'Moldavian', 'mon' : 'Mongolian', 'mos' : 'Mooré', 'mul' : 'Multiple languages', 'mun' : 'Munda (Other)', 'mus' : 'Creek', 'mwl' : 'Mirandese', 'mwr' : 'Marwari', 'myn' : 'Mayan languages', 'myv' : 'Erzya', 'nah' : 'Nahuatl', 'nai' : 'North American Indian (Other)', 'nap' : 'Neapolitan Italian', 'nau' : 'Nauru', 'nav' : 'Navajo', 'nbl' : 'Ndebele (South Africa)', 'nde' : 'Ndebele (Zimbabwe)', 'ndo' : 'Ndonga', 'nds' : 'Low German', 'nep' : 'Nepali', 'new' : 'Newari', 'nia' : 'Nias', 'nic' : 'Niger-Kordofanian (Other)', 'niu' : 'Niuean', 'nno' : 'Norwegian (Nynorsk)', 'nob' : 'Norwegian (Bokmål)', 'nog' : 'Nogai', 'non' : 'Old Norse', 'nor' : 'Norwegian', 'nqo' : 'N\'Ko', 'nso' : 'Northern Sotho', 'nub' : 'Nubian languages', 'nwc' : 'Newari, Old', 'nya' : 'Nyanja', 'nym' : 'Nyamwezi', 'nyn' : 'Nyankole', 'nyo' : 'Nyoro', 'nzi' : 'Nzima', 'oci' : 'Occitan (post-1500)', 'oji' : 'Ojibwa', 'ori' : 'Oriya', 'orm' : 'Oromo', 'osa' : 'Osage', 'oss' : 'Ossetic', 'ota' : 'Turkish, Ottoman', 'oto' : 'Otomian languages', 'paa' : 'Papuan (Other)', 'pag' : 'Pangasinan', 'pal' : 'Pahlavi', 'pam' : 'Pampanga', 'pan' : 'Panjabi', 'pap' : 'Papiamento', 'pau' : 'Palauan', 'peo' : 'Old Persian (ca. 600-400 B.C.)', 'per' : 'Persian', 'phi' : 'Philippine (Other)', 'phn' : 'Phoenician', 'pli' : 'Pali', 'pol' : 'Polish', 'pon' : 'Pohnpeian', 'por' : 'Portuguese', 'pra' : 'Prakrit languages', 'pro' : 'Provençal (to 1500)', 'pus' : 'Pushto', 'que' : 'Quechua', 'raj' : 'Rajasthani', 'rap' : 'Rapanui', 'rar' : 'Rarotongan', 'roa' : 'Romance (Other)', 'roh' : 'Raeto-Romance', 'rom' : 'Romani', 'rum' : 'Romanian', 'run' : 'Rundi', 'rup' : 'Aromanian', 'rus' : 'Russian', 'sad' : 'Sandawe', 'sag' : 'Sango (Ubangi Creole)', 'sah' : 'Yakut', 'sai' : 'South American Indian (Other)', 'sal' : 'Salishan languages', 'sam' : 'Samaritan Aramaic', 'san' : 'Sanskrit', 'sao' : 'Samoan', 'sas' : 'Sasak', 'sat' : 'Santali', 'scc' : 'Serbian', 'scn' : 'Sicilian Italian', 'sco' : 'Scots', 'scr' : 'Croatian', 'sel' : 'Selkup', 'sem' : 'Semitic (Other)', 'sga' : 'Irish, Old (to 1100)', 'sgn' : 'Sign languages', 'shn' : 'Shan', 'sho' : 'Shona', 'sid' : 'Sidamo', 'sin' : 'Sinhalese', 'sio' : 'Siouan (Other)', 'sit' : 'Sino-Tibetan (Other)', 'sla' : 'Slavic (Other)', 'slo' : 'Slovak', 'slv' : 'Slovenian', 'sma' : 'Southern Sami', 'sme' : 'Northern Sami', 'smi' : 'Sami', 'smj' : 'Lule Sami', 'smn' : 'Inari Sami', 'smo' : 'Samoan', 'sms' : 'Skolt Sami', 'sna' : 'Shona', 'snd' : 'Sindhi', 'snh' : 'Sinhalese', 'snk' : 'Soninke', 'sog' : 'Sogdian', 'som' : 'Somali', 'son' : 'Songhai', 'sot' : 'Sotho', 'spa' : 'Spanish', 'srd' : 'Sardinian', 'srn' : 'Sranan', 'srp' : 'Serbian', 'srr' : 'Serer', 'ssa' : 'Nilo-Saharan (Other)', 'sso' : 'Sotho', 'ssw' : 'Swazi', 'suk' : 'Sukuma', 'sun' : 'Sundanese', 'sus' : 'Susu', 'sux' : 'Sumerian', 'swa' : 'Swahili', 'swe' : 'Swedish', 'swz' : 'Swazi', 'syc' : 'Syriac', 'syr' : 'Syriac, Modern', 'tag' : 'Tagalog', 'tah' : 'Tahitian', 'tai' : 'Tai (Other)', 'taj' : 'Tajik', 'tam' : 'Tamil', 'tar' : 'Tatar', 'tat' : 'Tatar', 'tel' : 'Telugu', 'tem' : 'Temne', 'ter' : 'Terena', 'tet' : 'Tetum', 'tgk' : 'Tajik', 'tgl' : 'Tagalog', 'tha' : 'Thai', 'tib' : 'Tibetan', 'tig' : 'Tigré', 'tir' : 'Tigrinya', 'tiv' : 'Tiv', 'tkl' : 'Tokelauan', 'tlh' : 'Klingon (Artificial language)', 'tli' : 'Tlingit', 'tmh' : 'Tamashek', 'tog' : 'Tonga (Nyasa)', 'ton' : 'Tongan', 'tpi' : 'Tok Pisin', 'tru' : 'Truk', 'tsi' : 'Tsimshian', 'tsn' : 'Tswana', 'tso' : 'Tsonga', 'tsw' : 'Tswana', 'tuk' : 'Turkmen', 'tum' : 'Tumbuka', 'tup' : 'Tupi languages', 'tur' : 'Turkish', 'tut' : 'Altaic (Other)', 'tvl' : 'Tuvaluan', 'twi' : 'Twi', 'tyv' : 'Tuvinian', 'udm' : 'Udmurt', 'uga' : 'Ugaritic', 'uig' : 'Uighur', 'ukr' : 'Ukrainian', 'umb' : 'Umbundu', 'und' : 'Undetermined', 'urd' : 'Urdu', 'uzb' : 'Uzbek', 'vai' : 'Vai', 'ven' : 'Venda', 'vie' : 'Vietnamese', 'vol' : 'Volapük', 'vot' : 'Votic', 'wak' : 'Wakashan languages', 'wal' : 'Wolayta', 'war' : 'Waray', 'was' : 'Washoe', 'wel' : 'Welsh', 'wen' : 'Sorbian (Other)', 'wln' : 'Walloon', 'wol' : 'Wolof', 'xal' : 'Oirat', 'xho' : 'Xhosa', 'yao' : 'Yao (Africa)', 'yap' : 'Yapese', 'yid' : 'Yiddish', 'yor' : 'Yoruba', 'ypk' : 'Yupik languages', 'zap' : 'Zapotec', 'zbl' : 'Blissymbolics', 'zen' : 'Zenaga', 'zha' : 'Zhuang', 'znd' : 'Zande languages', 'zul' : 'Zulu', 'zun' : 'Zuni', 'zza' : 'Zaza'}
author_dict = {'abr' : 'abridger', 'acp' : 'art copyist', 'act' : 'actor', 'adi' : 'art director', 'adp' : 'adapter', 'aft' : 'author of afterword, colophon, etc.', 'anl' : 'analyst', 'anm' : 'animator', 'ann' : 'annotator', 'ant' : 'bibliographic antecedent', 'ape' : 'appellee', 'apl' : 'appellant', 'app' : 'applicant', 'aqt' : 'author in quotations or text abstracts', 'arc' : 'architect', 'ard' : 'artistic director', 'arr' : 'arranger', 'art' : 'artist', 'asg' : 'assignee', 'asn' : 'associated name', 'ato' : 'autographer', 'att' : 'attributed name', 'auc' : 'auctioneer', 'aud' : 'author of dialog', 'aui' : 'author of introduction, etc.', 'aus' : 'screenwriter', 'aut' : 'author', 'bdd' : 'binding designer', 'bjd' : 'bookjacket designer', 'bkd' : 'book designer', 'bkp' : 'book producer', 'blw' : 'blurb writer', 'bnd' : 'binder', 'bpd' : 'bookplate designer', 'brd' : 'broadcaster', 'brl' : 'braille embosser', 'bsl' : 'bookseller', 'cas' : 'caster', 'ccp' : 'conceptor', 'chr' : 'choreographer', 'clb' : 'contributor', 'cli' : 'client', 'cll' : 'calligrapher', 'clr' : 'colorist', 'clt' : 'collotyper', 'cmm' : 'commentator', 'cmp' : 'composer', 'cmt' : 'compositor', 'cnd' : 'conductor', 'cng' : 'cinematographer', 'cns' : 'censor', 'coe' : 'contestant-appellee', 'col' : 'collector', 'com' : 'compiler', 'con' : 'conservator', 'cor' : 'collection registrar', 'cos' : 'contestant', 'cot' : 'contestant-appellant', 'cou' : 'court governed', 'cov' : 'cover designer', 'cpc' : 'copyright claimant', 'cpe' : 'complainant-appellee', 'cph' : 'copyright holder', 'cpl' : 'complainant', 'cpt' : 'complainant-appellant', 'cre' : 'creator', 'crp' : 'correspondent', 'crr' : 'corrector', 'crt' : 'court reporter', 'csl' : 'consultant', 'csp' : 'consultant to a project', 'cst' : 'costume designer', 'ctb' : 'contributor', 'cte' : 'contestee-appellee', 'ctg' : 'cartographer', 'ctr' : 'contractor', 'cts' : 'contestee', 'ctt' : 'contestee-appellant', 'cur' : 'curator', 'cwt' : 'commentator for written text', 'dbp' : 'distribution place', 'dfd' : 'defendant', 'dfe' : 'defendant-appellee', 'dft' : 'defendant-appellant', 'dgg' : 'degree granting institution', 'dgs' : 'degree supervisor', 'dis' : 'dissertant', 'dln' : 'delineator', 'dnc' : 'dancer', 'dnr' : 'donor', 'dpc' : 'depicted', 'dpt' : 'depositor', 'drm' : 'draftsman', 'drt' : 'director', 'dsr' : 'designer', 'dst' : 'distributor', 'dtc' : 'data contributor', 'dte' : 'dedicatee', 'dtm' : 'data manager', 'dto' : 'dedicator', 'dub' : 'dubious author', 'edc' : 'editor of compilation', 'edm' : 'editor of moving image work', 'edt' : 'editor', 'ed' : 'editor', 'ед' : 'editor', 'egr' : 'engraver', 'elg' : 'electrician', 'elt' : 'electrotyper', 'eng' : 'engineer', 'enj' : 'enacting jurisdiction', 'etr' : 'etcher', 'evp' : 'event place', 'exp' : 'expert', 'fac' : 'facsimilist', 'fds' : 'film distributor', 'fld' : 'field director', 'flm' : 'film editor', 'fmd' : 'film director', 'fmk' : 'filmmaker', 'fmo' : 'former owner', 'fmp' : 'film producer', 'fnd' : 'funder', 'fpy' : 'first party', 'frg' : 'forger', 'gis' : 'geographic information specialist', 'grt' : 'artist', 'his' : 'host institution', 'hnr' : 'honoree', 'hst' : 'host', 'ill' : 'illustrator', 'ilu' : 'illuminator', 'ins' : 'inscriber', 'inv' : 'inventor', 'isb' : 'issuing body', 'itr' : 'instrumentalist', 'ive' : 'interviewee', 'ivr' : 'interviewer', 'jud' : 'judge', 'jug' : 'jurisdiction governed', 'lbr' : 'laboratory', 'lbt' : 'librettist', 'ldr' : 'laboratory director', 'led' : 'lead', 'lee' : 'libelee-appellee', 'lel' : 'libelee', 'len' : 'lender', 'let' : 'libelee-appellant', 'lgd' : 'lighting designer', 'lie' : 'libelant-appellee', 'lil' : 'libelant', 'lit' : 'libelant-appellant', 'lsa' : 'landscape architect', 'lse' : 'licensee', 'lso' : 'licensor', 'ltg' : 'lithographer', 'lyr' : 'lyricist', 'mcp' : 'music copyist', 'mdc' : 'metadata contact', 'med' : 'medium', 'mfp' : 'manufacture place', 'mfr' : 'manufacturer', 'mod' : 'moderator', 'mon' : 'monitor', 'mrb' : 'marbler', 'mrk' : 'markup editor', 'msd' : 'musical director', 'mte' : 'metal-engraver', 'mtk' : 'minute taker', 'mus' : 'musician', 'nrt' : 'narrator', 'opn' : 'opponent', 'org' : 'originator', 'orm' : 'organizer', 'osp' : 'onscreen presenter', 'oth' : 'other', 'own' : 'owner', 'pan' : 'panelist', 'pat' : 'patron', 'pbd' : 'publishing director', 'pbl' : 'publisher', 'pdr' : 'project director', 'pfr' : 'proofreader', 'pht' : 'photographer', 'plt' : 'platemaker', 'pma' : 'permitting agency', 'pmn' : 'production manager', 'pop' : 'printer of plates', 'ppm' : 'papermaker', 'ppt' : 'puppeteer', 'pra' : 'praeses', 'prc' : 'process contact', 'prd' : 'production personnel', 'pre' : 'presenter', 'prf' : 'performer', 'prg' : 'programmer', 'prm' : 'printmaker', 'prn' : 'production company', 'pro' : 'producer', 'prp' : 'production place', 'prs' : 'production designer', 'prt' : 'printer', 'prv' : 'provider', 'pta' : 'patent applicant', 'pte' : 'plaintiff-appellee', 'ptf' : 'plaintiff', 'pth' : 'patent holder', 'ptt' : 'plaintiff-appellant', 'pup' : 'publication place', 'rbr' : 'rubricator', 'rcd' : 'recordist', 'rce' : 'recording engineer', 'rcp' : 'addressee', 'rdd' : 'radio director', 'red' : 'redaktor', 'ren' : 'renderer', 'res' : 'researcher', 'rev' : 'reviewer', 'rpc' : 'radio producer', 'rps' : 'repository', 'rpt' : 'reporter', 'rpy' : 'responsible party', 'rse' : 'respondent-appellee', 'rsg' : 'restager', 'rsp' : 'respondent', 'rsr' : 'restorationist', 'rst' : 'respondent-appellant', 'rth' : 'research team head', 'rtm' : 'research team member', 'sad' : 'scientific advisor', 'sce' : 'scenarist', 'scl' : 'sculptor', 'scr' : 'scribe', 'sds' : 'sound designer', 'sec' : 'secretary', 'sgd' : 'stage director', 'sgn' : 'signer', 'sht' : 'supporting host', 'sll' : 'seller', 'sng' : 'singer', 'spk' : 'speaker', 'spn' : 'sponsor', 'spy' : 'second party', 'srv' : 'surveyor', 'std' : 'set designer', 'stg' : 'setting', 'stl' : 'storyteller', 'stm' : 'stage manager', 'stn' : 'standards body', 'str' : 'stereotyper', 'tcd' : 'technical director', 'tch' : 'teacher', 'ths' : 'thesis advisor', 'tld' : 'television director', 'tlp' : 'television producer', 'trc' : 'transcriber', 'trl' : 'translator', 'tr' : 'translator', 'tyd' : 'type designer', 'tyg' : 'typographer', 'uvp' : 'university place', 'vac' : 'voice actor', 'vdg' : 'videographer', 'voc' : 'singer', 'wac' : 'writer of added commentary', 'wal' : 'writer of added lyrics', 'wam' : 'writer of accompanying material', 'wat' : 'writer of added text', 'wdc' : 'woodcutter', 'wde' : 'wood engraver', 'win' : 'writer of introduction', 'wit' : 'witness', 'wpr' : 'writer of preface', 'wst' : 'writer of supplementary textual content'}

collection_dict = {}
collection_dict['monograph'] = {'coverage_depth' : 'ebook', 'oclc_collection_name' : 'Center for Research Libraries (CRL) eResources, Monographs', 'oclc_collection_id' : 'customer.93175.5'}
collection_dict['newspaper'] = {'coverage_depth' : 'fulltext', 'oclc_collection_name' : 'Center for Research Libraries (CRL) eResources, Newspapers', 'oclc_collection_id' : 'customer.93175.10'}
collection_dict['serial'] = {'coverage_depth' : 'fulltext', 'oclc_collection_name' : 'Center for Research Libraries (CRL) eResources, Serials', 'oclc_collection_id' : 'customer.93175.8'}

file_types_dict = {'csv' : '.csv', 'tsv' : '.tsv', 'excel' : '.xlsx', 'unicode text' : '.txt'}

file_types = [('csv file', '.csv'), ('tsv file', '.tsv'), ('excel file', '.xlsx'), ('unicode text', '.txt')]

record_source_dict = {}
record_source_dict['millennium'] = {'button' : 'Millennium', 'label' : 'Bib number:\t'}
record_source_dict['worldcat'] = {'button' : 'Worldcat', 'label' : 'OCLC number:\t'}
record_source_dict['folio'] = {'button' : 'Folio (uuid)', 'label' : 'Folio UUID:\t'}
record_source_dict['folio_oclc'] = {'button' : 'Folio (oclc)', 'label' : 'Folio OCLC number:\t'}
record_source_dict['ddsnext'] = {'button' : 'DDSnext', 'label' : ''}
record_source_dict['eastview'] = {'button' : 'EastView', 'label' : ''}
record_source_dict['icon'] = {'button' : 'ICON', 'label' : ''}
record_source_dict['recap'] = {'button' : 'ReCAP', 'label' : ''}
record_source_dict['papr'] = {'button' : 'PAPR', 'label' : ''}

#Simple Excel writer class
class excel_writer:
    def __init__(self, workbook, filename, name):
        self.row = 1
        self.column = 1
        self.workbook = workbook
        self.filename = filename
        self.workbook_sheet = workbook.active
        self.workbook_sheet.title = name
    #Write row to workbook.
    def writerow(self, row_content):
        for i in range(1, len(row_content) + 1):
            if row_content[i - 1] == 'None' or row_content[i - 1] is None:
                row_content[i - 1] = ''
            self.workbook_sheet.cell(row = self.row, column = i, value=row_content[i - 1])
        self.row += 1
    #Save workbook.
    def save(self):
        self.workbook.save(filename = self.filename)

class main_configuration:
    def __init__(self):
        self.config_folder = os.path.join(os.path.join(os.path.join(os.path.join(os.path.join('C:\\Users', os.getlogin()), 'AppData'), 'Local'), 'CRL'), 'Metadata')
        self.config = configparser.ConfigParser()
        self.config_file = os.path.join(self.config_folder, 'metadata_extraction_config.ini')
        self.config_data = {}
        self.read_config_file()
    #Read configuration file.  Create one if none exist.
    def read_config_file(self):
        check_or_create_dir(self.config_folder)
        #Create a blank file if none exists
        if not os.path.isfile(self.config_file):
            self.write_config_file()
        self.config.read(self.config_file)
    #Write to configuration file.
    def write_config_file(self):
        with open(self.config_file, 'w') as config_out:
            self.config.write(config_out)
    #Check to see if a section with a given name exist.
    def section_exist(self, section):
        if section in self.config:
            return True
        return False
    #Create a new section if one with a given name does not exist.
    def add_section(self, section):
        if not self.section_exist(section):
            self.config[section] = None
            self.write_config_file()
    #Add template location.
    def add_template_location(self, section, file_type):
        if not self.section_exist(section):
            self.config[section] = {}
        if file_type not in self.config[section]:
            self.config[section][file_type] = section + '.' + file_type
            self.config[self.config[section][file_type]] = {}
            self.write_config_file()
    #Modifies the file location in the config file.
    def modify_file_location(self, section, file_type=None, folder_name=None, file_name=None, file_extention=None, file_location=None):
        #Gets file data from directly.
        if file_location is None:
            file_dict = OrderedDict({'folder_name' : folder_name, 'file_name' : file_name, 'file_extention' : file_extention})
        #Get file data from given file location.
        else:
            file_dict = OrderedDict({'folder_name' : os.path.dirname(file_location), 'file_name' : os.path.splitext(os.path.basename(file_location))[0], 'file_extention' : os.path.splitext(os.path.basename(file_location))[1]})
        #Add file data to configuration file.
        if not self.section_exist(section) or file_type not in self.config[section]:
            self.add_template_location(section, file_type)
            self.config[self.config[section][file_type]] = file_dict
            self.write_config_file()
        else:
            self.config[self.config[section][file_type]] = file_dict
            self.write_config_file()
    #Gets file location.
    def get_file_location(self, section, file_type, filename_modifier = None):
        if self.section_exist(section) and self.config[section][file_type] in self.config:
            if filename_modifier is not None:
                return os.path.join(self.config[self.config[section][file_type]]['folder_name'], self.config[self.config[section][file_type]]['file_name'] + filename_modifier + self.config[self.config[section][file_type]]['file_extention'])
            else:
                return os.path.join(self.config[self.config[section][file_type]]['folder_name'], self.config[self.config[section][file_type]]['file_name'] + self.config[self.config[section][file_type]]['file_extention'])
        return None

class api_configuration:
    def __init__(self):
        self.config_folder = os.path.join(os.path.join(os.path.join(os.path.join(os.path.join('C:\\Users', os.getlogin()), 'AppData'), 'Local'), 'CRL'), 'Metadata')
        check_or_create_dir(self.config_folder)
        if not os.path.isfile(self.config_folder):
            self.initial_config()
            self.add_section('API KEYS')
#            self.config['API KEYS'] = {'tremayne' : 'dJWOW13v3Dto56AkOoQr6iyrQcLiTYt8kMV0vVAMGNo4LCjj67QmEu34sgMHKJVZ59oz7h3CmTooSAyd', 'chad' : 'OFMCzlRWkYW1JeJoSUCHRWPHsAbNKzbYcXjVnt9M5OYkPIdAkSwkEQVSUQuMUqOsmXUV4rMmnkDcxFBS'}
            self.add_section('Preferred API Key')
#            self.config['Preferred API Key'] = {'preferred' : 'chad'}
            self.write_config_file()
        else:
            self.initial_config()
    #Create Worldcat API configuration file.
    def initial_config(self):
        self.config = configparser.ConfigParser()
        self.config_file = os.path.join(self.config_folder, 'api_manager.ini')
        self.config_data = {}
        self.read_config_file()
    #Read Worldcat API configuration file.  Create one if none exist.
    def read_config_file(self):
        check_or_create_dir(self.config_folder)
        # create a blank file if none exists
        if not os.path.isfile(self.config_file):
            self.write_config_file()
        self.config.read(self.config_file)
    #Updates the Worldcat API configuration file with the current configuration
    def write_config_file(self):
        with open(self.config_file, 'w') as config_out:
            self.config.write(config_out)
    #Check to see if a section with a given name exist.
    def section_exist(self, section):
        if section in self.config:
            return True
        return False
    #Create a new section if one with a given name does not exist.
    def add_section(self, section):
        if not self.section_exist(section):
            self.config[section] = {}
            self.write_config_file()
    #Get preferred Worldcat API key.
    def get_apikey(self):
        return self.config['API KEYS'][self.config['Preferred API Key']['preferred']]

api_keys = api_configuration()
folio_config = configuration()


#Export dialog class
class export_dialog(tkinter.Toplevel):
    def __init__(self, parent, title = None):
        tkinter.Toplevel.__init__(self, parent)
        self.transient(parent)
        if title:
            self.title(title)
        self.parent = parent
        self.result = None
        self.title('Export')
        self.resizable(width='FALSE', height='FALSE')
        
        self.config = self.parent.config
        
        self.entry_descriptive_metadata_text = tkinter.StringVar()
        self.entry_title_metadata_text = tkinter.StringVar()
        self.entry_kbart_metadata_text = tkinter.StringVar()
        
        body = tkinter.Frame(self)
        self.initial_focus = self.body(body)
        self.grab_set()
        if not self.initial_focus:
            self.initial_focus = self
        self.protocol("WM_DELETE_WINDOW", self.teminate)
        self.geometry("+%d+%d" % (parent.winfo_rootx()+50, parent.winfo_rooty()+50))
        self.initial_focus.focus_set()
        self.wait_window(self)
    #Creates the body of the export dialog box.
    def body(self, master):
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)
        self.columnconfigure(1, weight = 10)
        self.rowconfigure(4, weight = 10)
        self.top_frame = tkinter.Frame(self)
        self.top_frame.grid(row=0, sticky="nwe")
        
        self.descriptive_metadata_label = tkinter.Label(self.top_frame, text='Descriptive Metadata File:\t', font=font.Font(size=self.parent.size))
        self.title_metadata_label = tkinter.Label(self.top_frame, text='Title Metadata File:\t', font=font.Font(size=self.parent.size))
        self.kbart_metadata_label = tkinter.Label(self.top_frame, text='Kbart Metadata File:\t', font=font.Font(size=self.parent.size))
        
        self.descriptive_metadata_button = tkinter.Button(self.top_frame, text='Browse', font=font.Font(size=self.parent.size), command=self.open_descriptive_metadata_file)
        self.title_metadata_button = tkinter.Button(self.top_frame, text='Browse', font=font.Font(size=self.parent.size), command=self.open_title_metadata_file)
        self.kbart_metadata_button = tkinter.Button(self.top_frame, text='Browse', font=font.Font(size=self.parent.size), command=self.open_kbart_metadata_file)
        
        self.save_button = tkinter.Button(self.top_frame, text='Export', font=font.Font(size=self.parent.size), command=self.export)
        self.close_button = tkinter.Button(self.top_frame, text='Close', font=font.Font(size=self.parent.size), command=self.destroy)
        
        self.entry_descriptive_metadata = tkinter.Entry(self.top_frame, textvariable=self.entry_descriptive_metadata_text, font=font.Font(size=self.parent.size))
        self.entry_title_metadata = tkinter.Entry(self.top_frame, textvariable=self.entry_title_metadata_text, font=font.Font(size=self.parent.size))
        self.entry_kbart_metadata = tkinter.Entry(self.top_frame, textvariable=self.entry_kbart_metadata_text, font=font.Font(size=self.parent.size))
        
        self.descriptive_metadata_label.grid(row=0, column=0, sticky='N'+'W')
        self.title_metadata_label.grid(row=1, column=0, sticky='N'+'W')
        self.kbart_metadata_label.grid(row=2, column=0, sticky='N'+'W')
        
        self.entry_descriptive_metadata.grid(row=0, column=1, columnspan=20, sticky='N'+'E'+'W')
        self.entry_title_metadata.grid(row=1, column=1, columnspan=20, sticky='N'+'E'+'W')
        self.entry_kbart_metadata.grid(row=2, column=1, columnspan=20, sticky='N'+'E'+'W')
        
        self.descriptive_metadata_button.grid(row=0, column=21, sticky='N'+'E'+'W')
        self.title_metadata_button.grid(row=1, column=21, sticky='N'+'E'+'W')
        self.kbart_metadata_button.grid(row=2, column=21, sticky='N'+'E'+'W')
        
        self.save_button.grid(row=4, column=20)
        self.close_button.grid(row=4, column=21)
        self.grid_rowconfigure(0, weight=1)
        col_count, row_count = self.top_frame.grid_size()
        for col in range(col_count):
            self.top_frame.grid_columnconfigure(col, minsize=20)
        for row in range(row_count):
            self.top_frame.grid_rowconfigure(row, minsize=20)
    #Returns true if all row entries are empty.  Returns false otherwise.
    def check_row(self, spreadsheet, row):
        verify_empty = True
        for item in spreadsheet[row]:
            if spreadsheet[row][item].get() != '':
                verify_empty = False
                break
        return verify_empty
    #Export data to files.
    def export(self):
        #Exports descriptive metadata data to file if the descriptive metadata Entry is not empty.
        if self.entry_descriptive_metadata_text.get() != '':
            if self.config.config['file_locations.descriptive_metadata']['file_extention'] == '.xlsx':
                descriptive_metadata_file = openpyxl.Workbook()
                dds_descriptive_writer = excel_writer(descriptive_metadata_file, self.config.get_file_location('file_locations', 'descriptive_metadata'), 'descriptive_metadata')
            elif self.config.config['file_locations.descriptive_metadata']['file_extention'] == '.txt':
                descriptive_metadata_file = open(self.config.get_file_location('file_locations', 'descriptive_metadata'),'wt',encoding='utf-16-le', newline='')
                dds_descriptive_writer = csv.writer(descriptive_metadata_file, delimiter='\t', dialect='excel-tab')
            else:
                descriptive_metadata_file = open(self.config.get_file_location('file_locations', 'descriptive_metadata'),'wt',encoding='utf8', newline='')
                if self.config.config['file_locations.descriptive_metadata']['file_extention'] == '.tsv':
                    dds_descriptive_writer = csv.writer(descriptive_metadata_file, delimiter='\t', dialect='excel')
                else:
                    dds_descriptive_writer = csv.writer(descriptive_metadata_file, dialect='excel')
            for row in self.parent.template_spreadsheet['descriptive_metadata']:
                if not self.check_row(self.parent.template_spreadsheet['descriptive_metadata'], row):
                    dds_descriptive_writer.writerow([self.parent.template_spreadsheet['descriptive_metadata'][row]['title_uuid'].get(), self.parent.template_spreadsheet['descriptive_metadata'][row]['field'].get(), self.parent.template_spreadsheet['descriptive_metadata'][row]['value'].get()])
            if self.config.config['file_locations.descriptive_metadata']['file_extention'] == '.xlsx':
                dds_descriptive_writer.save()
            descriptive_metadata_file.close()
        #Exports title metadata data to file if the title metadata Entry is not empty.
        if self.entry_title_metadata_text.get() != '':
            if self.config.config['file_locations.title_metadata']['file_extention'] == '.xlsx':
                title_metadata_file = openpyxl.Workbook()
                dds_title_writer = excel_writer(title_metadata_file, self.config.get_file_location('file_locations', 'title_metadata'), 'title_metadata')
            elif self.config.config['file_locations.title_metadata']['file_extention'] == '.txt':
                title_metadata_file = open(self.config.get_file_location('file_locations', 'title_metadata'),'wt',encoding='utf-16-le', newline='')
                dds_title_writer = csv.writer(title_metadata_file, delimiter='\t', dialect='excel-tab')
            else:
                title_metadata_file = open(self.config.get_file_location('file_locations', 'title_metadata'),'wt',encoding='utf8', newline='')
                if self.config.config['file_locations.title_metadata']['file_extention'] == '.tsv':
                    dds_title_writer = csv.writer(title_metadata_file, delimiter='\t', dialect='excel')
                else:
                    dds_title_writer = csv.writer(title_metadata_file, dialect='excel')
            for row in self.parent.template_spreadsheet['title_metadata']:
                if not self.check_row(self.parent.template_spreadsheet['title_metadata'], row):
                    dds_title_writer.writerow([self.parent.template_spreadsheet['title_metadata'][row]['title_name'].get(), self.parent.template_spreadsheet['title_metadata'][row]['title_uuid'].get(), self.parent.template_spreadsheet['title_metadata'][row]['title_material_type'].get(), self.parent.template_spreadsheet['title_metadata'][row]['title_format'].get(), self.parent.template_spreadsheet['title_metadata'][row]['title_oclc'].get(), self.parent.template_spreadsheet['title_metadata'][row]['title_digital_holding_range'].get(), self.parent.template_spreadsheet['title_metadata'][row]['title_resolution'].get(), self.parent.template_spreadsheet['title_metadata'][row]['title_color_depth'].get(), self.parent.template_spreadsheet['title_metadata'][row]['title_location_code'].get(), self.parent.template_spreadsheet['title_metadata'][row]['title_catalog_link'].get(), self.parent.template_spreadsheet['title_metadata'][row]['title_external_link'].get(), self.parent.template_spreadsheet['title_metadata'][row]['item_name'].get(), self.parent.template_spreadsheet['title_metadata'][row]['item_uuid'].get(), self.parent.template_spreadsheet['title_metadata'][row]['item_pub_year'].get(), self.parent.template_spreadsheet['title_metadata'][row]['item_pub_month'].get(), self.parent.template_spreadsheet['title_metadata'][row]['item_pub_day'].get(), self.parent.template_spreadsheet['title_metadata'][row]['item_group'].get(), self.parent.template_spreadsheet['title_metadata'][row]['item_ocr'].get(), self.parent.template_spreadsheet['title_metadata'][row]['item_coordinated_ocr'].get(), self.parent.template_spreadsheet['title_metadata'][row]['item_pdf'].get(), self.parent.template_spreadsheet['title_metadata'][row]['item_access'].get()])
            if self.config.config['file_locations.title_metadata']['file_extention'] == '.xlsx':
                dds_title_writer.save()
            title_metadata_file.close()
        #Exports kbart metadata data to file if the kbart metadata Entry is not empty.
        if self.entry_kbart_metadata_text.get() != '':
            kbart_file_dict = {}
            #Sorts kbart into kbart_file_dict using oclc_collection_id.
            for row in self.parent.template_spreadsheet['kbart_metadata']:
                if row != 'header':
                    if self.parent.template_spreadsheet['kbart_metadata'][row]['oclc_collection_id'].get() == 'customer.93175.5':
                        kbart_file_dict['customer.93175.5'] = {'type' : 'monograph', 'worldshare_writer' : None}
                    elif self.parent.template_spreadsheet['kbart_metadata'][row]['oclc_collection_id'].get() == 'customer.93175.10':
                        kbart_file_dict['customer.93175.10'] = {'type' : 'newspaper', 'worldshare_writer' : None}
                    elif self.parent.template_spreadsheet['kbart_metadata'][row]['oclc_collection_id'].get() == 'customer.93175.8':
                        kbart_file_dict['customer.93175.8'] = {'type' : 'serial', 'worldshare_writer' : None}
                    elif self.parent.template_spreadsheet['kbart_metadata'][row]['oclc_collection_id'].get() != '' and self.parent.template_spreadsheet['kbart_metadata'][row]['oclc_collection_id'].get() is not None:
                        kbart_file_dict[self.parent.template_spreadsheet['kbart_metadata'][row]['oclc_collection_id'].get()] = {'type' : '', 'worldshare_writer' : None}
            if kbart_file_dict == {}:
                if self.config.config['file_locations.kbart_metadata']['file_extention'] == '.xlsx':
                    kbart_metadata_file = openpyxl.Workbook()
                    worldshare_writer = excel_writer(kbart_metadata_file, self.config.get_file_location('file_locations', 'kbart_metadata'), 'kbart_metadata')
                elif self.config.config['file_locations.kbart_metadata']['file_extention'] == '.txt':
                    kbart_metadata_file = open(self.config.get_file_location('file_locations', 'kbart_metadata'),'wt',encoding='utf-16-le', newline='')
                    worldshare_writer = csv.writer(kbart_metadata_file, delimiter='\t', dialect='excel-tab')
                else:
                    kbart_metadata_file = open(self.config.get_file_location('file_locations', 'kbart_metadata'),'wt',encoding='utf8', newline='')
                    if self.config.config['file_locations.kbart_metadata']['file_extention'] == '.tsv':
                        worldshare_writer = csv.writer(kbart_metadata_file, delimiter='\t', dialect='excel')
                    else:
                        worldshare_writer = csv.writer(kbart_metadata_file, dialect='excel')
            else:
                if self.config.config['file_locations.kbart_metadata']['file_extention'] == '.xlsx':
                    for key in kbart_file_dict:
                        kbart_metadata_file = openpyxl.Workbook()
                        kbart_file_dict[key]['worldshare_writer'] = excel_writer(kbart_metadata_file, self.config.get_file_location('file_locations', 'kbart_metadata', filename_modifier = '_' + kbart_file_dict[key]['type']), 'kbart_metadata')
                elif self.config.config['file_locations.kbart_metadata']['file_extention'] == '.txt':
                    for key in kbart_file_dict:
                        kbart_metadata_file = open(self.config.get_file_location('file_locations', 'kbart_metadata', filename_modifier = '_' + kbart_file_dict[key]['type']),'wt',encoding='utf-16-le', newline='')
                        kbart_file_dict[key]['worldshare_writer'] = csv.writer(kbart_metadata_file, delimiter='\t', dialect='excel-tab')
                else:
                    for key in kbart_file_dict:
                        kbart_metadata_file = open(self.config.get_file_location('file_locations', 'kbart_metadata', filename_modifier = '_' + kbart_file_dict[key]['type']),'wt',encoding='utf8', newline='')
                        if self.config.config['file_locations.kbart_metadata']['file_extention'] == '.tsv':
                            kbart_file_dict[key]['worldshare_writer'] = csv.writer(kbart_metadata_file, delimiter='\t', dialect='excel')
                        else:
                            kbart_file_dict[key]['worldshare_writer'] = csv.writer(kbart_metadata_file, dialect='excel')
            if kbart_file_dict == {}:
                for row in self.parent.template_spreadsheet['kbart_metadata']:
                    worldshare_writer.writerow([self.parent.template_spreadsheet['kbart_metadata'][row]['publication_title'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['print_identifier'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['online_identifier'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['date_first_issue_online'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['num_first_vol_online'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['num_first_issue_online'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['date_last_issue_online'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['num_last_vol_online'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['num_last_issue_online'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['title_url'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['first_author'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['title_id'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['embargo_info'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['coverage_depth'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['coverage_notes'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['publisher_name'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['location'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['title_notes'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['staff_notes'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['vendor_id'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['oclc_collection_name'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['oclc_collection_id'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['oclc_entry_id'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['oclc_linkscheme'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['oclc_number'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['action'].get()])
            else:
                for row in self.parent.template_spreadsheet['kbart_metadata']:
                    if row == 'header':
                        for key in kbart_file_dict:
                            kbart_file_dict[key]['worldshare_writer'].writerow([self.parent.template_spreadsheet['kbart_metadata'][row]['publication_title'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['print_identifier'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['online_identifier'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['date_first_issue_online'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['num_first_vol_online'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['num_first_issue_online'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['date_last_issue_online'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['num_last_vol_online'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['num_last_issue_online'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['title_url'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['first_author'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['title_id'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['embargo_info'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['coverage_depth'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['coverage_notes'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['publisher_name'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['location'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['title_notes'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['staff_notes'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['vendor_id'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['oclc_collection_name'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['oclc_collection_id'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['oclc_entry_id'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['oclc_linkscheme'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['oclc_number'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['action'].get()])
                    elif not self.check_row(self.parent.template_spreadsheet['kbart_metadata'], row):
                        kbart_file_dict[self.parent.template_spreadsheet['kbart_metadata'][row]['oclc_collection_id'].get()]['worldshare_writer'].writerow([self.parent.template_spreadsheet['kbart_metadata'][row]['publication_title'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['print_identifier'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['online_identifier'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['date_first_issue_online'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['num_first_vol_online'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['num_first_issue_online'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['date_last_issue_online'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['num_last_vol_online'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['num_last_issue_online'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['title_url'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['first_author'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['title_id'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['embargo_info'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['coverage_depth'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['coverage_notes'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['publisher_name'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['location'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['title_notes'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['staff_notes'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['vendor_id'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['oclc_collection_name'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['oclc_collection_id'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['oclc_entry_id'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['oclc_linkscheme'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['oclc_number'].get(), self.parent.template_spreadsheet['kbart_metadata'][row]['action'].get()])

            if self.config.config['file_locations.kbart_metadata']['file_extention'] == '.xlsx':
                if kbart_file_dict == {}:
                    worldshare_writer.save()
                for key in kbart_file_dict:
                    kbart_file_dict[key]['worldshare_writer'].save()
            kbart_metadata_file.close()
        self.teminate()
    #Opens file dialog
    def open_file(self, initialdir, confirmoverwrite, defaultextension, filetypes, initialfile, title):
        return asksaveasfilename(initialdir = initialdir, confirmoverwrite = confirmoverwrite, defaultextension = defaultextension, filetypes = filetypes, initialfile = initialfile, title = title)
    #Opens descriptive metadata file dialog, and prints it to descriptive metadata entry text field
    def open_descriptive_metadata_file(self):
        file_location = self.open_file(initialdir = self.config.config['file_locations.descriptive_metadata']['folder_name'], confirmoverwrite = True, defaultextension = self.config.config['file_locations.descriptive_metadata']['file_extention'], filetypes = file_types, initialfile = self.config.config['file_locations.descriptive_metadata']['file_name'], title = 'Descriptive metadata')
        if file_location:
            self.entry_descriptive_metadata_text.set(file_location)
            self.config.modify_file_location('file_locations', 'descriptive_metadata', file_location=file_location)
    #Opens title metadata file dialog, and prints it to title metadata entry text field
    def open_title_metadata_file(self):
        file_location = self.open_file(initialdir = self.config.config['file_locations.title_metadata']['folder_name'], confirmoverwrite = True, defaultextension = self.config.config['file_locations.title_metadata']['file_extention'], filetypes = file_types, initialfile = self.config.config['file_locations.title_metadata']['file_name'], title = 'Title metadata')
        if file_location:
            self.entry_title_metadata_text.set(file_location)
            self.config.modify_file_location('file_locations', 'title_metadata', file_location=file_location)
    #Opens kbart metadata file dialog, and prints it to kbart metadata entry text field
    def open_kbart_metadata_file(self):
        file_location = self.open_file(initialdir = self.config.config['file_locations.kbart_metadata']['folder_name'], confirmoverwrite = True, defaultextension = self.config.config['file_locations.kbart_metadata']['file_extention'], filetypes = [('unicode text', '.txt'), ('csv file', '.csv'), ('tsv file', '.tsv'), ('excel file', '.xlsx')], initialfile = self.config.config['file_locations.kbart_metadata']['file_name'], title = 'Kbart metadata')
        if file_location:
            self.entry_kbart_metadata_text.set(file_location)
            self.config.modify_file_location('file_locations', 'kbart_metadata', file_location=file_location)
    #Sets focus to parent application and closes dialog
    def teminate(self):
        self.parent.focus_set()
        self.destroy()

#Folio settings dialog class
class folio_settings_dialog(tkinter.Toplevel):
    def __init__(self, parent, title = None):
        tkinter.Toplevel.__init__(self, parent)
        self.transient(parent)
        if title:
            self.title(title)
        self.parent = parent
        self.result = None
        self.title('Folio API')
        self.resizable(width='FALSE', height='FALSE')
        
        self.config = self.parent.config
        
        self.entry_username_text = tkinter.StringVar()
        self.entry_password_text = tkinter.StringVar()
        self.entry_okapi_text = tkinter.StringVar()
        self.entry_tenant_text = tkinter.StringVar()
        
        self.password = config.config['data']['password']
        self.visible = False
        
        self.entry_username_text.set(config.config['data']['username'])
        self.entry_password_text.set(self.password)
        self.entry_okapi_text.set(config.config['data']['okapi_url'])
        self.entry_tenant_text.set(config.config['data']['tenant'])
        
        body = tkinter.Frame(self)
        self.initial_focus = self.body(body)
        self.grab_set()
        if not self.initial_focus:
            self.initial_focus = self
        self.protocol("WM_DELETE_WINDOW", self.teminate)
        self.geometry("+%d+%d" % (parent.winfo_rootx()+50, parent.winfo_rooty()+50))
        self.initial_focus.focus_set()
        self.wait_window(self)
    #Creates the body of the export dialog box.
    def body(self, master):
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)
        self.columnconfigure(1, weight = 10)
        self.rowconfigure(4, weight = 10)
        self.top_frame = tkinter.Frame(self)
        self.top_frame.grid(row=0, sticky="nwe")
        
        self.username_label = tkinter.Label(self.top_frame, text='Username:\t', font=font.Font(size=self.parent.size))
        self.password_label = tkinter.Label(self.top_frame, text='Password:\t', font=font.Font(size=self.parent.size))
        self.okapi_label = tkinter.Label(self.top_frame, text='OKAPI url:\t', font=font.Font(size=self.parent.size))
        self.tenant_label = tkinter.Label(self.top_frame, text='Tenant:\t', font=font.Font(size=self.parent.size))
        
        self.password_button = tkinter.Button(self.top_frame, text='Show', font=font.Font(size=self.parent.size), command=self.toogle_visiblity)
        
        self.apply_button = tkinter.Button(self.top_frame, text='Apply', font=font.Font(size=self.parent.size), command=self.apply)
        self.save_button = tkinter.Button(self.top_frame, text='Save', font=font.Font(size=self.parent.size), command=self.save)
        self.close_button = tkinter.Button(self.top_frame, text='Close', font=font.Font(size=self.parent.size), command=self.destroy)
        
        self.entry_username = tkinter.Entry(self.top_frame, textvariable=self.entry_username_text, font=font.Font(size=self.parent.size))
        self.entry_password = tkinter.Entry(self.top_frame, textvariable=self.entry_password_text, font=font.Font(size=self.parent.size), show="*")
        self.entry_okapi = tkinter.Entry(self.top_frame, textvariable=self.entry_okapi_text, font=font.Font(size=self.parent.size))
        self.entry_tenant = tkinter.Entry(self.top_frame, textvariable=self.entry_tenant_text, font=font.Font(size=self.parent.size))
        
        self.username_label.grid(row=0, column=0, sticky='N'+'W')
        self.password_label.grid(row=1, column=0, sticky='N'+'W')
        self.okapi_label.grid(row=2, column=0, sticky='N'+'W')
        self.tenant_label.grid(row=3, column=0, sticky='N'+'W')
        
        self.entry_username.grid(row=0, column=1, columnspan=21, sticky='N'+'E'+'W')
        self.entry_password.grid(row=1, column=1, columnspan=21, sticky='N'+'E'+'W')
        self.entry_okapi.grid(row=2, column=1, columnspan=21, sticky='N'+'E'+'W')
        self.entry_tenant.grid(row=3, column=1, columnspan=21, sticky='N'+'E'+'W')
        
        self.password_button.grid(row=1, column=22, sticky='N'+'E'+'W')
        
        self.apply_button.grid(row=4, column=20)
        self.save_button.grid(row=4, column=21)
        self.close_button.grid(row=4, column=22)
        self.grid_rowconfigure(0, weight=1)
        col_count, row_count = self.top_frame.grid_size()
        for col in range(col_count):
            self.top_frame.grid_columnconfigure(col, minsize=20)
        for row in range(row_count):
            self.top_frame.grid_rowconfigure(row, minsize=20)
    #Toogle password visiblity.
    def toogle_visiblity(self):
        self.visible = not self.visible
        if self.visible:
            self.entry_password['show'] = ''
            self.password_button['text'] = 'Hide'
        elif not self.visible:
            self.entry_password['show'] = '*'
            self.password_button['text'] = 'Show'
    #Apply Folio api settings.
    def apply(self):
        if self.entry_username_text.get() != '':
            folio_config.config['data']['okapi_url'] = self.entry_okapi_text.get()
        if self.entry_password_text.get() != '':
            folio_config.config['data']['tenant'] = self.entry_tenant_text.get()
        if self.entry_username_text.get() != '':
            folio_config.config['data']['username'] = self.entry_username_text.get()
        if self.entry_password_text.get() != '':
            folio_config.config['data']['password'] = self.entry_password_text.get()
        auth(okapi_url = folio_config.config['data']['okapi_url'], tenant = folio_config.config['data']['tenant'], username = folio_config.config['data']['username'], password = folio_config.config['data']['password'])
        folio_config.config['data']['okapi_token'] = get_token()
    #Saves Folio api setting to file.
    def save(self):
        if self.entry_username_text.get() != '':
            folio_config.config['data']['okapi_url'] = self.entry_okapi_text.get()
        if self.entry_password_text.get() != '':
            folio_config.config['data']['tenant'] = self.entry_tenant_text.get()
        if self.entry_username_text.get() != '':
            folio_config.config['data']['username'] = self.entry_username_text.get()
        if self.entry_password_text.get() != '':
            folio_config.config['data']['password'] = self.entry_password_text.get()
        auth(okapi_url = folio_config.config['data']['okapi_url'], tenant = folio_config.config['data']['tenant'], username = folio_config.config['data']['username'], password = folio_config.config['data']['password'])
        folio_config.config['data']['okapi_token'] = get_token()
        folio_config.write_config_file()
        self.teminate()
    #Sets focus to parent application and closes dialog
    def teminate(self):
        self.parent.focus_set()
        self.destroy()

#Convert list to dictionary
def convert_to_dict(input_list, header=False):
    output_dict = {}
    row = 0
    for value in input_list:
        if value == 'header':
            output_dict['header'] = value
            header=False
        else:
            output_dict[row] = value
        row += 1
    return output_dict

#Remove all carriage returns from text.
def remove_returns(text):
    while re.match('(.+)(?:\r)(.*$)', text):
        text = re.match('(.+)(?:\r)(.*$)', text).group(1) + re.match('(.+)(?:\r)(.*$)', text).group(2)
    return text

#Removes start/end whitespace characters.
def remove_whitespace(text):
    if re.match('(.+)(?:\s+$)', text):
        text = re.match('(.+)(?:\s+$)', text).group(1)
    if re.match('(?:\s+)(.+$)', text):
        text = re.match('(?:\s+)(.+$)', text).group(1)
    return text

#Main application class.
class Application(tkinter.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.bib_num = ''
        self.uuid = ''
        self.size = 11
        self.master.title('Metadata Extraction')
        self.master.geometry('1000x700')
        self.grid(row=0, sticky='N'+'S'+'E'+'W')
        self.master.grid_rowconfigure(0, weight = 1)
        self.master.grid_columnconfigure(0, weight = 1)
        self.screen_text = tkinter.StringVar()
        self.entry_catalog_id_text = tkinter.StringVar()
        self.entry_uuid_text = tkinter.StringVar()
        self.entry_page_text = tkinter.StringVar()
        self.number_of_pages = tkinter.StringVar()
        self.number_of_pages.set(0)
        self.catalog_id_text = tkinter.StringVar()
        self.catalog_id_text.set('Folio UUID:\t')
        self.catalog_id_button_text = tkinter.StringVar()
        self.catalog_id_button_text.set('Folio (uuid)')
        self.record_source = 'folio'
        self.file_types = 'csv'
        self.input_values = []
        self.import_type = 'batch'
        self.records = {}
        self.pages = {}
        self.spreadsheet = {}
        self.template_spreadsheet = {}
        self.template_dict = {}
        self.windows_id_dict = {}
        self.windows = ttk.Notebook(self)
        self.screen_frame = tkinter.Frame(self.windows)
        self.spreadsheet_frame = tkinter.Frame(self.windows)
        self.create_menu()
        self.create_record_viewer()
        self.create_spreadsheet()
        self.windows.add(self.spreadsheet_frame, text='Input')
        self.windows_id_dict['Input'] = 0
        self.windows.add(self.screen_frame, text='Records')
        self.windows_id_dict['Records'] = 1
        
        self.windows.grid(row=0, column=0, rowspan=8, columnspan=8, sticky='nswe')
        
        self.title_metadata = {}
        self.descriptive_metadata = {}
        self.kbart_metadata = {}
        
        global output_folder
        
        self.config = main_configuration()
        self.config.modify_file_location('file_locations', file_type='descriptive_metadata', folder_name=output_folder, file_name='ddsnext_descriptive_metadata_' + get_date(), file_extention=file_types_dict[self.file_types])
        self.config.modify_file_location('file_locations', file_type='title_metadata', folder_name=output_folder, file_name='ddsnext_title_metadata_' + get_date(), file_extention=file_types_dict[self.file_types])
        self.config.modify_file_location('file_locations', file_type='kbart_metadata', folder_name=output_folder, file_name='worldshare_metadata_' + get_date(), file_extention='unicode text')
        
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)
        self.entry_input_text = tkinter.StringVar()
        self.entry_output_text = tkinter.StringVar()
        self.entry_error_text = tkinter.StringVar()
        self.current_record = None
    #Creates main menu
    def create_menu(self):
        self.menubar = tkinter.Menu(self.master)
        self.file_menu = tkinter.Menu(self.menubar, tearoff=0)
        self.file_menu.add_command(label='Run', font=font.Font(size=self.size), command=self.run)
        self.file_menu.add_command(label='Export', font=font.Font(size=self.size), state = 'disabled', command=self.export)
        self.file_menu.add_command(label='Quit', font=font.Font(size=self.size), command=self.teminate)
        self.menubar.add_cascade(label='File', font=font.Font(size=self.size), menu=self.file_menu)
        
        self.setting_menu = tkinter.Menu(self.menubar, tearoff=0)
        self.setting_menu.add_command(label='Folio API', font=font.Font(size=self.size), command=self.set_folio_settings)
        self.setting_menu.add_command(label='Worldcat API', font=font.Font(size=self.size), state = 'disabled', command=self.set_folio_settings)
        self.menubar.add_cascade(label='Setting', font=font.Font(size=self.size), menu=self.setting_menu)
        #Displays the menu
        self.master.config(menu=self.menubar)
    #Creates Marc record viewer tab.
    def create_record_viewer(self):
        self.screen_frame.grid(row=0, column=0, rowspan=8, columnspan=8, sticky='nswe')
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)
        self.screen_frame.grid_rowconfigure(1, weight=1)
        self.screen_frame.grid_columnconfigure(0, weight=1)
        self.screen_frame_outer = tkinter.Frame(self.screen_frame, background='black')
        #Sets up the screen canvas to enable scrolling.
        self.screen_canvas = tkinter.Canvas(self.screen_frame_outer, background='black', highlightthickness=0)
        self.screen_frame_inner = tkinter.Frame(self.screen_canvas, background='black')
        #Sets up the vertical and horizontal scrollbars.
        self.screen_scrollbar_vertical=tkinter.Scrollbar(self.screen_frame_outer, orient='vertical', command=self.screen_canvas.yview)
        self.screen_scrollbar_horizontal=tkinter.Scrollbar(self.screen_frame_outer, orient='horizontal', command=self.screen_canvas.xview)
        self.scrollbar_visible = False
        self.screen_label = tkinter.Label(self.screen_frame_inner, background='black', textvariable=self.screen_text, font=font.Font(size=self.size), fg='white', anchor='nw', justify=tkinter.LEFT)
        self.screen_label.grid(row=0, column=0, columnspan=8, rowspan=6, sticky='nswe')
        self.screen_frame_inner.grid_rowconfigure(0, weight=1, minsize=461)
        self.screen_frame_inner.grid_columnconfigure(0, weight=1, minsize=620)
        #Adds space to the frame to prevent the vertical scrollbar from cutting off text.
        self.screen_frame_inner.grid_columnconfigure(9, minsize=15)
        self.screen_frame_inner.grid_rowconfigure(8, weight=1)
        self.screen_frame_inner.grid_columnconfigure(6, weight=1)
        
        #Finishes setting up the main canvas by connecting the scrollbar.
        self.screen_canvas.configure(yscrollcommand=self.screen_scrollbar_vertical.set, xscrollcommand=self.screen_scrollbar_horizontal.set)
        self.canvas_frame = self.screen_canvas.create_window((0,0), window=self.screen_frame_inner, anchor='nw')
        
        #Binds the canvases to the configure event.
        self.screen_frame_outer.bind('<Configure>', self.set_up_canvas)
        self.screen_frame_outer.bind('<<change_page>>', self.set_up_records)
        
        self.screen_frame_outer.grid(row=1, column=0, columnspan=8, sticky='nswe', padx=6)
        self.screen_frame_outer.grid_rowconfigure(0, weight=1)
        self.screen_frame_outer.grid_columnconfigure(0, weight=1)
        
        self.screen_canvas.grid(row=0, column=0, columnspan=8, sticky='nswe', padx=6)
        
        self.page_frame = tkinter.Frame(self.screen_frame)
        self.first_page_button = tkinter.Button(self.page_frame, text='<<', font=font.Font(size=8, weight='bold'), command=partial(self.change_record_page_navigation, first=True))
        self.previous_page_button = tkinter.Button(self.page_frame, text='<', font=font.Font(size=8, weight='bold'), command=partial(self.change_record_page_navigation, change=-1))
        self.entry_page = tkinter.Entry(self.page_frame, textvariable=self.entry_page_text, font=font.Font(size=self.size))
        self.page_divider = tkinter.Label(self.page_frame, text=' / ', font=font.Font(size=self.size))
        self.page_label = tkinter.Label(self.page_frame, textvariable=self.number_of_pages, font=font.Font(size=self.size))
        self.next_page_button = tkinter.Button(self.page_frame, text='>', font=font.Font(size=8, weight='bold'), command=partial(self.change_record_page_navigation, change=1))
        self.last_page_button = tkinter.Button(self.page_frame, text='>>', font=font.Font(size=8, weight='bold'), command=partial(self.change_record_page_navigation, last=True))
        self.page_frame.grid(row=2, column=0)
        
        self.first_page_button.grid(row=0, column=1)
        self.previous_page_button.grid(row=0, column=2)
        self.entry_page.grid(row=0, column=3)
        self.page_divider.grid(row=0, column=4)
        self.page_label.grid(row=0, column=5)
        self.next_page_button.grid(row=0, column=6)
        self.last_page_button.grid(row=0, column=7)
        self.grid_rowconfigure(2, weight=1)
        self.grid_columnconfigure(3, weight=1)
        self.grid_columnconfigure(6, weight=1)
        
        self.entry_page.bind('<Return>', self.change_record_page_event_handler)
    #Creates main spreadsheet for input data.
    def create_spreadsheet(self):
        self.spreadsheet_frame_main = tkinter.Frame(self.spreadsheet_frame)
        self.spreadsheet_frame.grid_rowconfigure(0, weight=1)
        
        self.spreadsheet_frame.grid_columnconfigure(0, weight=1)
        
        #Binds the canvases to the configure event.
        self.spreadsheet_frame_main.bind('<Configure>', self.set_up_canvas)
        
        #Sets up the screen canvas to enable scrolling.
        self.spreadsheet_screen_canvas = tkinter.Canvas(self.spreadsheet_frame_main, highlightthickness=0)
        self.spreadsheet_frame_inner = tkinter.Frame(self.spreadsheet_screen_canvas)
        
        #Sets up the vertical and horizontal scrollbars.
        self.spreadsheet_scrollbar_vertical=tkinter.Scrollbar(self.spreadsheet_frame_main, orient='vertical', command=self.spreadsheet_screen_canvas.yview)
        self.spreadsheet_scrollbar_horizontal=tkinter.Scrollbar(self.spreadsheet_frame_main, orient='horizontal', command=self.spreadsheet_screen_canvas.xview)
        self.spreadsheet_scrollbar_vertical.grid(row=0, column=8, sticky='ns')
        self.spreadsheet_scrollbar_horizontal.grid(row=1, column=0, columnspan=8, sticky='we')
        
        #Finishes setting up the main canvas by connecting the scrollbar and adding the frame containing hot folders information.
        self.spreadsheet_screen_canvas.configure(yscrollcommand=self.spreadsheet_scrollbar_vertical.set, xscrollcommand=self.spreadsheet_scrollbar_horizontal.set)
        
        self.canvas_frame = self.spreadsheet_screen_canvas.create_window((0,0), window=self.spreadsheet_frame_inner, anchor='nw', tags='self.spreadsheet_frame_inner')
        
        self.spreadsheet_frame_main.grid_rowconfigure(0, weight=1)
        self.spreadsheet_frame_main.grid_columnconfigure(0, weight=1)
        
        self.spreadsheet_screen_canvas.grid(row=0, column=0, columnspan=8, sticky='nswe')
        
        self.spreadsheet_catalog_id_label = tkinter.Label(self.spreadsheet_frame_inner, textvariable=self.catalog_id_text, font=font.Font(size=self.size))
        self.spreadsheet_uuid_label = tkinter.Label(self.spreadsheet_frame_inner, text='DDSnext UUID:\t', font=font.Font(size=self.size))
        self.spreadsheet_file_location_label = tkinter.Label(self.spreadsheet_frame_inner, text='File Location:\t', font=font.Font(size=self.size))
        #Sets up button to add collection info to spreadsheet.
        self.spreadsheet_collection_button = tkinter.Menubutton(self.spreadsheet_frame_inner, text='Collection:\t', font=font.Font(size=self.size), relief=tkinter.RAISED, anchor='w')
        self.spreadsheet_collection_button.grid()
        self.spreadsheet_collection_button.menu = tkinter.Menu(self.spreadsheet_collection_button, tearoff=0)
        self.spreadsheet_collection_button['menu'] =  self.spreadsheet_collection_button.menu
        self.spreadsheet_collection_button.menu.add_command(label='Monograph', font=font.Font(size=self.size), command=partial(self.change_collection, 'monograph'))
        self.spreadsheet_collection_button.menu.add_command(label='Newspaper', font=font.Font(size=self.size), command=partial(self.change_collection, 'newspaper'))
        self.spreadsheet_collection_button.menu.add_command(label='Serial', font=font.Font(size=self.size), command=partial(self.change_collection, 'serial'))



        self.spreadsheet_catalog_id_label.grid(row=0, column=0, sticky='w')
        self.spreadsheet_uuid_label.grid(row=0, column=4, sticky='w')
        self.spreadsheet_file_location_label.grid(row=0, column=8, sticky='w')
        
        self.spreadsheet_collection_button.grid(row=0, column=16, sticky='nswe')
        
        
        self.spreadsheet_bottom_frame = tkinter.Frame(self.spreadsheet_frame)
        self.spreadsheet_run_button = tkinter.Button(self.spreadsheet_bottom_frame, text='Run', font=font.Font(size=self.size), command=self.run)
        self.spreadsheet_clear_button = tkinter.Button(self.spreadsheet_bottom_frame, text='Clear', font=font.Font(size=self.size), state = 'disabled', command=self.clear)
        
        self.spreadsheet_clear_options_button = tkinter.Menubutton(self.spreadsheet_bottom_frame, text='^', font=font.Font(size=5, weight='bold'), relief=tkinter.RAISED)
        self.spreadsheet_clear_options_button.menu = tkinter.Menu(self.spreadsheet_clear_options_button, tearoff=0)
        
        self.clear_value = tkinter.StringVar()
        self.clear_value.set('all')
        self.spreadsheet_clear_options_button['menu'] =  self.spreadsheet_clear_options_button.menu
        self.spreadsheet_clear_options_button.menu.add_radiobutton(label='Entries and Records', variable=self.clear_value, value='all', font=font.Font(size=self.size), command = self.get_clear_state)
        self.spreadsheet_clear_options_button.menu.add_radiobutton(label='Entries', variable=self.clear_value, value='entries', font=font.Font(size=self.size), command = self.get_clear_state)
        self.spreadsheet_clear_options_button.menu.add_radiobutton(label='Records', variable=self.clear_value, value='records', font=font.Font(size=self.size), command = self.get_clear_state)
        
        #Sets up Button to switch between Millennium and Worldcat.
        self.spreadsheet_catalog_id_button = tkinter.Menubutton(self.spreadsheet_bottom_frame, textvariable=self.catalog_id_button_text, font=font.Font(size=self.size), relief=tkinter.RAISED)
        self.spreadsheet_catalog_id_button.grid()
        self.spreadsheet_catalog_id_button.menu = tkinter.Menu(self.spreadsheet_catalog_id_button, tearoff=0)
        self.spreadsheet_catalog_id_button['menu'] =  self.spreadsheet_catalog_id_button.menu
        self.spreadsheet_catalog_id_button.menu.add_command(label='Folio (uuid)', font=font.Font(size=self.size), command=partial(self.change_record_source, 'folio'))
        self.spreadsheet_catalog_id_button.menu.add_command(label='Folio (oclc)', font=font.Font(size=self.size), command=partial(self.change_record_source, 'folio_oclc'))
        self.spreadsheet_catalog_id_button.menu.add_command(label='Worldcat', font=font.Font(size=self.size), command=partial(self.change_record_source, 'worldcat'))
        
        
        self.spreadsheet_frame_main.grid(row=0, column=0, sticky='nswe')
        self.spreadsheet_bottom_frame.grid(row=1, column=0)
        self.spreadsheet_run_button.grid(row=1, column=1)
        self.spreadsheet_clear_button.grid(row=1, column=2)
        self.spreadsheet_clear_options_button.grid(row=1, column=3, sticky='nw')
        self.spreadsheet_catalog_id_button.grid(row=1, column=0, sticky='nswe')
        
        self.set_up_spreadsheet()
        self.clipboard_content = ''
    
    #Changes record source.
    def change_record_source(self, record_source):
        self.catalog_id_button_text.set(record_source_dict[record_source]['button'])
        self.catalog_id_text.set(record_source_dict[record_source]['label'])
        self.record_source = record_source
    #Changes collection and fills entries in input spreadsheet.
    def change_collection(self, collection):
        for row in self.spreadsheet:
            self.spreadsheet[row]['collection'][0].set(collection)
    #Sets settings
    def set_folio_settings(self):
        folio_settings_dialog(self)
    
    #Gets Marc records from user provided data then creates title_metadata, descriptive_metadata, and kbart_metadata spreadsheets.
    def run(self):
        self.screen_text.set('')
        self.input_values = []
        for row in self.spreadsheet:
            if self.spreadsheet[row]['cat_id'][0].get() is not None and self.spreadsheet[row]['cat_id'][0].get() != '' and self.spreadsheet[row]['ddsnext_uuid'][0].get() is not None and self.spreadsheet[row]['ddsnext_uuid'][0].get() != '':
                self.input_values.append([remove_whitespace(remove_returns(self.spreadsheet[row]['cat_id'][0].get())), remove_whitespace(remove_returns(self.spreadsheet[row]['ddsnext_uuid'][0].get())), remove_whitespace(remove_returns(self.spreadsheet[row]['collection'][0].get().lower()))])
        self.process_records()
        self.pages = {}
        page_num = 0
        if self.records != {}:
            for key in list(self.records):
                page_num += 1
                self.pages[page_num] = key
            self.screen_text.set(self.records[self.pages[1]])
            self.entry_page_text.set(1)
            self.number_of_pages.set(page_num)
            self.current_record = [1, self.pages[1]]
            self.windows.select(1)
            self.spreadsheet_clear_button['state'] = 'normal'
            self.file_menu.entryconfig('Export', state= 'normal')
            self.update()
            self.screen_frame_outer.event_generate('<<change_page>>', when='tail')
    
    #Change record from data from application's navigation tools.
    def change_record_page_navigation(self, change=0, last=False, first=False):
        if first:
            self.entry_page_text.set(1)
            self.change_record_page()
        elif last:
           self.entry_page_text.set(self.number_of_pages.get())
           self.change_record_page()
        else:
            page_num = int(self.entry_page_text.get()) + change
            if page_num > 0 and page_num <= int(self.number_of_pages.get()):
                self.entry_page_text.set(page_num)
                self.change_record_page()
    #Change record by page.
    def change_record_page(self):
        self.screen_text.set('')
        page_num = int(self.entry_page.get())
        if self.pages != {} and self.records != {}:
            if page_num in self.pages:
                self.screen_text.set(self.records[self.pages[page_num]])
                self.entry_page_text.set(page_num)
            else:
                self.entry_page_text.set(self.current_record[0])
            self.update()
            self.screen_frame_outer.event_generate('<<change_page>>', when='tail')
    #Returns true if all spreadsheet entries are empty.  Returns false otherwise.
    def check_entries_empty(self):
        verify_empty = True
        for row in self.spreadsheet:
            for key in self.spreadsheet[row]:
                if self.spreadsheet[row][key][0].get() != '':
                    verify_empty = False
                    break
            if verify_empty == False:
                break
                verify_empty = False
                break
        return verify_empty
    #
    def retry_get_clear_state(self, event):
        while self.attempt < 10:
            self.update()
            self.attempt += 1
            if not self.check_entries_empty():
                self.get_clear_state()
                break
        self.attempt = 1
        event.widget.unbind('<<retry_get_clear_state>>')
        self.get_clear_state()
    
    #Event handler that determines if the clear button should be disabled or enabled.
    def check_entries_empty_handler(self, event):
        self.attempt = 1
        event.widget.bind('<<retry_get_clear_state>>', self.retry_get_clear_state)
        event.widget.event_generate('<<retry_get_clear_state>>', when='tail')
    #Determines if the clear button should be disabled or enabled.
    def get_clear_state(self):
        verify_empty = True
        #Checks if the spreadsheet entries are empty if clear_value is 'all' or 'entries'.
        if self.clear_value.get() == 'all' or self.clear_value.get() == 'entries':
            verify_empty = self.check_entries_empty()
        #Checks if there are any marc records if clear_value is 'all' or 'records'.
        if self.clear_value.get() == 'all' or self.clear_value.get() == 'records':
            if self.records != {}:
                verify_empty = False
        #If no valid check returned any values, disables the check button.
        if verify_empty:
            self.spreadsheet_clear_button['state'] = 'disabled'
        #If any valid check returned any values, enables the check button.
        else:
            self.spreadsheet_clear_button['state'] = 'normal'
    
    #Event handler that determines if the clear button should be disabled or enabled.
    def get_clear_state_handler(self, event):
        self.get_clear_state()
    
    #Clears screen, records, and spreadsheet.
    def clear(self):
        #Reset the screen text, record, and page info if clear_value is 'all' or 'records'.
        if self.clear_value.get() == 'all' or self.clear_value.get() == 'records':
            self.screen_text.set('')
            self.records = {}
            self.pages = {}
            self.entry_page_text.set(0)
            self.number_of_pages.set(0)
            self.screen_scrollbar_vertical.grid_forget()
            self.screen_scrollbar_horizontal.grid_forget()
            self.scrollbar_visible = False
        #Reset the spreadsheet entries if clear_value is 'all' or 'entries'.
        if self.clear_value.get() == 'all' or self.clear_value.get() == 'entries':
            for row in self.spreadsheet:
                for key in self.spreadsheet[row]:
                    self.spreadsheet[row][key][0].set('')
        #Disables the check button.
        self.spreadsheet_clear_button['state'] = 'disabled'
        self.file_menu.entryconfig('Export', state= 'disabled')
    
    #Clears template spreadsheet.
    def clear_template(self, template_name):
        for row in self.template_spreadsheet[template_name]:
            for item in self.template_spreadsheet[template_name][row]:
                if type(self.template_spreadsheet[template_name][row][item]) is tkinter.Entry:
                    self.set_entry(self.template_spreadsheet[template_name][row][item])
    #Empties template spreadsheet.
    def empty_template(self, template_name):
        for row in self.template_spreadsheet[template_name]:
            for item in self.template_spreadsheet[template_name][row]:
                self.template_spreadsheet[template_name][row][item].destroy()
            if row == 'header':
                row = 0
            self.template_dict[template_name]['template_spreadsheet_frame'].grid_rowconfigure(row, weight=0, minsize=0)
        self.template_spreadsheet[template_name] = {}
    #Sets entry value even if disabled.
    def set_entry(self, given_entry, text=''):
        disable = False
        if given_entry['state'] == 'disabled':
            given_entry['state'] = 'normal'
            disable = True
        given_entry.delete(0, tkinter.END)
        given_entry.insert(0, text)
        if disable:
            given_entry['state'] = 'disabled'
    
    #Sets up the screen and spreadsheet canvas for scrolling.
    def set_up_canvas(self, event):
        width = event.width - 1
        self.screen_canvas.configure(scrollregion=self.screen_canvas.bbox('all'))
        self.spreadsheet_screen_canvas.configure(scrollregion=self.spreadsheet_screen_canvas.bbox('all'))
    
    #Sets up the template canvases for scrolling.
    def set_up_template_canvas(self, event):
        for template_name in self.template_dict:
            self.update()
            self.template_dict[template_name]['template_screen_canvas'].configure(scrollregion=self.template_dict[template_name]['template_screen_canvas'].bbox('all'))
    
    #Sets up the screen canvas for scrolling.
    def set_up_records(self, event):
        self.screen_canvas.configure(scrollregion=self.screen_canvas.bbox('all'))
        if not self.scrollbar_visible:
            self.screen_scrollbar_vertical.grid(row=0, column=2, rowspan=2, sticky='ns')
            self.screen_scrollbar_horizontal.grid(row=1, column=0, sticky='we')
            self.scrollbar_visible = True
        self.screen_canvas.yview_moveto('0')
        self.screen_canvas.xview_moveto('0')
        
    #Opens folder dialog
    def open_folder_dialog(self, initialdir, title):
        return askdirectory(initialdir = initialdir, title = title, mustexist=tkinter.TRUE)
    #Opens input file dialog, and prints it to input entry text field
    def get_folder_location(self, spreadsheet_row):
        if os.path.exists(spreadsheet_row['folder_location'][0].get()):
            file_location = self.open_folder_dialog(initialdir = spreadsheet_row['folder_location'][0].get(), title = spreadsheet_row['ddsnext_uuid'][0].get() + ' scan location')
        else:
            file_location = self.open_folder_dialog(initialdir = os.path.dirname(os.path.realpath('__file__')), title = spreadsheet_row['ddsnext_uuid'][0].get() + ' scan location')
        if file_location:
            spreadsheet_row['folder_location'][0].set(file_location)
    #Sets up the spreadsheet.
    def set_up_spreadsheet(self):
        for row in range(1, 51):
            self.spreadsheet[row] = {'cat_id' : [tkinter.StringVar(), None], 'ddsnext_uuid' : [tkinter.StringVar(), None], 'folder_location' : [tkinter.StringVar(), None], 'folder_location_button' : [tkinter.StringVar(), None], 'collection' : [tkinter.StringVar(), None]}
            self.spreadsheet[row]['cat_id'][1] = tkinter.Entry(self.spreadsheet_frame_inner, textvariable=self.spreadsheet[row]['cat_id'][0], font=font.Font(size=self.size))
            self.spreadsheet[row]['ddsnext_uuid'][1] = tkinter.Entry(self.spreadsheet_frame_inner, textvariable=self.spreadsheet[row]['ddsnext_uuid'][0], font=font.Font(size=self.size))
            self.spreadsheet[row]['folder_location'][1] = tkinter.Entry(self.spreadsheet_frame_inner, textvariable=self.spreadsheet[row]['folder_location'][0], font=font.Font(size=self.size))
            self.spreadsheet[row]['folder_location_button'][1] = tkinter.Button(self.spreadsheet_frame_inner, text='Browse', font=font.Font(size=self.size), command=partial(self.get_folder_location, self.spreadsheet[row]))
            self.spreadsheet[row]['collection'][1] = tkinter.Entry(self.spreadsheet_frame_inner, textvariable=self.spreadsheet[row]['collection'][0], font=font.Font(size=self.size))
            self.spreadsheet[row]['cat_id'][1].grid(row=row, column=0, columnspan=4, sticky='nswe')
            self.spreadsheet[row]['ddsnext_uuid'][1].grid(row=row, column=4, columnspan=4, sticky='nswe')
            self.spreadsheet[row]['folder_location'][1].grid(row=row, column=8, columnspan=4, sticky='nswe')
            self.spreadsheet[row]['folder_location_button'][1].grid(row=row, column=12, columnspan=4, sticky='nswe')
            self.spreadsheet[row]['collection'][1].grid(row=row, column=13, columnspan=4, sticky='nswe')
            self.spreadsheet_frame_inner.grid_rowconfigure(row, weight=1, minsize=25)
            
            self.spreadsheet[row]['cat_id'][1].bind('<FocusIn>', self.enter)
            self.spreadsheet[row]['ddsnext_uuid'][1].bind('<FocusIn>', self.enter)
            self.spreadsheet[row]['folder_location'][1].bind('<FocusIn>', self.enter)
            self.spreadsheet[row]['collection'][1].bind('<FocusIn>', self.enter)
            self.spreadsheet[row]['cat_id'][1].bind('<FocusOut>', self.get_clear_state_handler)
            self.spreadsheet[row]['ddsnext_uuid'][1].bind('<FocusOut>', self.get_clear_state_handler)
            self.spreadsheet[row]['folder_location'][1].bind('<FocusOut>', self.get_clear_state_handler)
            self.spreadsheet[row]['collection'][1].bind('<FocusOut>', self.get_clear_state_handler)
        self.spreadsheet_frame_inner.grid_columnconfigure(0, weight=1, minsize=330)
        self.spreadsheet_frame_inner.grid_columnconfigure(4, weight=1, minsize=330)
        self.spreadsheet_frame_inner.grid_columnconfigure(8, weight=1, minsize=330)
        self.spreadsheet_frame_inner.grid_columnconfigure(12, weight=1)
        self.spreadsheet_frame_inner.grid_columnconfigure(16, weight=1, minsize=330)
    
    #Sets up the spreadsheet for given template.
    def set_up_template(self, template_name, data):
        #Creates template window and frame if not the window dictionary.
        if not template_name in self.windows_id_dict:
            #Creates the template frame.
            template_frame = tkinter.Frame(self.windows)
            template_frame.grid(row = 0, column = 0, sticky='nswe')
            template_frame.grid_columnconfigure(0, weight = 1)
            template_frame.grid_rowconfigure(0, weight = 1)
            template_frame.bind('<Configure>', self.set_up_template_canvas)
            template_frame.bind('<<update_template_canvas>>', self.set_up_template_canvas)
            
            #Creates the frame containing the spreadsheet canvas.
            template_spreadsheet_frame_outer = tkinter.Frame(template_frame)
            template_spreadsheet_frame_outer.grid_rowconfigure(0, weight = 1)
            template_spreadsheet_frame_outer.grid_columnconfigure(0, weight = 1)
            template_spreadsheet_frame_outer.grid(row = 0, column = 0, sticky='nswe')
            
            #Creates the bottom frame of the template frame.
            template_bottom_frame = tkinter.Frame(template_frame)
            template_bottom_frame.grid(row = 2, column = 0)
            
            #Creates the (outer) canvas for the spreadsheet frame.
            template_screen_canvas = tkinter.Canvas(template_spreadsheet_frame_outer, highlightthickness = 0)
            template_screen_canvas.grid(row = 0, column = 0, sticky='nswe')
            
            #Creates the spreadsheet frame of the template frame.
            template_spreadsheet_frame = tkinter.Frame(template_screen_canvas)
            
            #Sets up the vertical and horizontal scrollbars.
            spreadsheet_scrollbar_vertical = tkinter.Scrollbar(template_spreadsheet_frame_outer, orient = 'vertical', command = template_screen_canvas.yview)
            spreadsheet_scrollbar_horizontal = tkinter.Scrollbar(template_spreadsheet_frame_outer, orient = 'horizontal', command = template_screen_canvas.xview)
            spreadsheet_scrollbar_vertical.grid(row = 0, column = 2, sticky = 'ns')
            spreadsheet_scrollbar_horizontal.grid(row = 1, column = 0, sticky = 'we')
            
            
            #Finishes setting up the template canvas by connecting the scrollbar.
            template_screen_canvas.configure(yscrollcommand = spreadsheet_scrollbar_vertical.set, xscrollcommand = spreadsheet_scrollbar_horizontal.set)
            template_screen_canvas.create_window((0,0), window = template_spreadsheet_frame, anchor='nw', tags = template_name)
            
            #
            self.template_dict[template_name] = {'template_frame' : template_frame, 'template_spreadsheet_frame' : template_spreadsheet_frame, 'template_bottom_frame' : template_bottom_frame, 'template_screen_canvas' : template_screen_canvas}
            
            self.template_spreadsheet[template_name] = {}
            self.template_spreadsheet[template_name]['header'] = {}
            
            self.windows.add(self.template_dict[template_name]['template_frame'], text=template_name)
            self.windows_id_dict[template_name] = len(self.windows_id_dict)
            
        else:
            self.empty_template(template_name)
            self.update()
            self.template_spreadsheet[template_name]['header'] = {}
        
        self.template_dict[template_name]['template_spreadsheet_frame'].grid_columnconfigure(0, weight=1)
        
        row = 0
        item_row = 0
        #Creates header row in spreedsheet.
        column=0
        for data_item in data['header']:
            self.template_spreadsheet[template_name]['header'][data_item] = tkinter.Entry(self.template_dict[template_name]['template_spreadsheet_frame'], font=font.Font(size=self.size))
            to_print = ''
            if data['header'][data_item] is not None:
                to_print = data['header'][data_item]
            self.set_entry(self.template_spreadsheet[template_name]['header'][data_item], text = to_print)
            self.template_spreadsheet[template_name]['header'][data_item].grid(row=0, column=column, sticky='nswe')
            
            if template_name == 'descriptive_metadata':
                self.template_dict[template_name]['template_spreadsheet_frame'].grid_columnconfigure(column, weight=1, minsize=326)
            else:
                self.template_dict[template_name]['template_spreadsheet_frame'].grid_columnconfigure(column, weight=1, minsize=240)
            self.template_dict[template_name]['template_spreadsheet_frame'].grid_rowconfigure(row, weight=1, minsize=25)
            self.template_spreadsheet[template_name]['header'][data_item]['state'] = 'disabled'
            column += 1
        row += 1
        #Creates remaining rows in spreedsheet.
        for key in data:
            if key != 'header':
                column = 0
                for data_item in data[key]:
                    if row not in self.template_spreadsheet[template_name]:
                        self.template_spreadsheet[template_name][row] = {}
                    self.template_spreadsheet[template_name][row][data_item] = tkinter.Entry(self.template_dict[template_name]['template_spreadsheet_frame'], font=font.Font(size=self.size))
                    to_print = ''
                    if data[key][data_item] is not None:
                        to_print = data[key][data_item]
                    self.set_entry(self.template_spreadsheet[template_name][row][data_item], text = to_print)
                    self.template_spreadsheet[template_name][row][data_item].grid(row = row, column = column, sticky='nswe')
                    column += 1
                self.template_dict[template_name]['template_spreadsheet_frame'].grid_rowconfigure(row, weight=1, minsize=25)
                row += 1
        last_row = self.get_template_last_row(template_name)
        #Adds an additional row if there are more than 24 rows to ensure the last row with data is visible.
        if last_row < 24:
            for current_row in range(last_row, 24):
                self.add_template_row(template_name, last_row=current_row)
        template_run_button = tkinter.Button(self.template_dict[template_name]['template_bottom_frame'], text='Add row', font=font.Font(size=self.size), command=partial(self.add_template_row, template_name))
        template_clear_button = tkinter.Button(self.template_dict[template_name]['template_bottom_frame'], text='Clear', font=font.Font(size=self.size), command=partial(self.clear_template, template_name))
        
        template_run_button.grid(row=0, column=1)
        template_clear_button.grid(row=0, column=2)
        self.update()
    #Gets the row number of the last row of given template spreadsheet.
    def get_template_last_row(self, template_name):
        last_row = 0
        for row in self.template_spreadsheet[template_name]:
            if type(row) is int and row > last_row:
                last_row = row
        return last_row
    #Adds row to given template spreadsheet.
    def add_template_row(self, template_name, last_row=None):
        if last_row is None:
            last_row = 0
            for row in self.template_spreadsheet[template_name]:
                if type(row) is int and row > last_row:
                    last_row = row
        last_row += 1
        column = 0
        for key in self.template_spreadsheet[template_name]['header']:
            if column == 0:
                self.template_spreadsheet[template_name][last_row] = {}
            self.template_spreadsheet[template_name][last_row][key] = tkinter.Entry(self.template_dict[template_name]['template_spreadsheet_frame'], font=font.Font(size=self.size))
            self.template_spreadsheet[template_name][last_row][key].grid(row = last_row, column = column, sticky='nswe')
            column += 1
        
        self.template_dict[template_name]['template_spreadsheet_frame'].grid_rowconfigure(last_row, weight=1, minsize=25)
        
        self.template_dict[template_name]['template_frame'].event_generate('<<update_template_canvas>>', when='tail')
    
    def get_collection_data(self, template_name, row, collection):
        if 'coverage_depth' in self.template_spreadsheet[template_name][row]:
            self.set_entry(self.template_spreadsheet[template_name][row]['coverage_depth'], text=collection_dict[collection]['coverage_depth'])
        else:
            self.set_entry(self.template_spreadsheet[template_name][row][14], text=collection_dict[collection]['coverage_depth'])
        if 'oclc_collection_name' in self.template_spreadsheet[template_name][row]:
            self.set_entry(self.template_spreadsheet[template_name][row]['oclc_collection_name'], text=collection_dict[collection]['oclc_collection_name'])
        else:
            self.set_entry(self.template_spreadsheet[template_name][row][21], text=collection_dict[collection]['oclc_collection_name'])
        if 'oclc_collection_id' in self.template_spreadsheet[template_name][row]:
            self.set_entry(self.template_spreadsheet[template_name][row]['oclc_collection_id'], text=collection_dict[collection]['oclc_collection_id'])
        else:
            self.set_entry(self.template_spreadsheet[template_name][row][22], text=collection_dict[collection]['oclc_collection_id'])
        
    #Event handler that changes page if the page entry is the focus.
    def change_record_page_event_handler(self, event):
        if self.entry_page.focus_get():
            self.entry_page_text.set(self.entry_page.get())
            self.change_record_page()
    #When entering widget, binds check_clipboard to 'Ctrl-v'.
    def enter(self, event):
        event.widget.bind('<Control-KeyPress-v>', self.check_clipboard)
        event.widget.bind('<KeyPress>', self.check_entries_empty_handler)
    #Restores Windows clipboard.
    def restore_clipboard(self, event):
        attempt = 1
        while True:
            try:
                win32clipboard.OpenClipboard()
                win32clipboard.EmptyClipboard()
                win32clipboard.SetClipboardText(self.clipboard_content, win32clipboard.CF_TEXT)
                win32clipboard.CloseClipboard()
                event.widget.unbind('<<restore_clipboard>>')
                break
            except pywintypes.error:
                if pywintypes.error != 5 or attempt == 5:
                    break
                attempt += 1
                self.after(attempt * attempt * 10)
    #Parses clipboard text.  Tabs to next cell.  Next line to next row, same column.
    #Works for dynamic number of columns.
    def check_clipboard(self, event):
        attempt = 1
        while True:
            try:
                win32clipboard.OpenClipboard()
                self.clipboard_content = ''
                if win32clipboard.IsClipboardFormatAvailable(win32clipboard.CF_TEXT):
                    self.clipboard_content = win32clipboard.GetClipboardData()
                win32clipboard.EmptyClipboard()
                win32clipboard.CloseClipboard()
                self.clipboard_clear()
                current_widget = event.widget
                for match in re.finditer('([^\n]+)(\n|$)', self.clipboard_content, re.MULTILINE):
                    match_text = match.group(1)
                    tab_match = re.match('([^\t]+)(?:\t)(.+$)', match.group(1), re.MULTILINE)
                    number_of_tabs = 0
                    if tab_match:
                        while tab_match:
                            if type(current_widget) is not tkinter.Entry:
                                break
                            self.set_entry(current_widget, text = tab_match.group(1))
                            current_widget = current_widget.tk_focusNext()
                            match_text = tab_match.group(2)
                            tab_match = re.match('([^\t]+)(?:\t)(.+$)', match_text, re.MULTILINE)
                            number_of_tabs += 1
                        else:
                            if type(current_widget) is tkinter.Entry:
                                match_text = remove_returns(match_text)
                                self.set_entry(current_widget, text = match_text)
                                if re.match('(\n|$)' , match.group(2), re.MULTILINE):
                                    for tab in range(len(self.spreadsheet[1]) - number_of_tabs):
                                        current_widget = current_widget.tk_focusNext()
                    elif type(current_widget) is tkinter.Entry:
                        match_text = remove_returns(match_text)
                        self.set_entry(current_widget, text = match_text)
                        if re.match('(\n|$)' , match.group(2)):
                            for tab in range(len(self.spreadsheet[1])):
                                current_widget = current_widget.tk_focusNext()
                self.get_clear_state()
                current_widget.bind('<<restore_clipboard>>', self.restore_clipboard)
                current_widget.event_generate('<<restore_clipboard>>', when='tail')
                current_widget.focus()
                break
            except pywintypes.error:
                if pywintypes.error != 5 or attempt == 5:
                    break
                attempt += 1
                self.after(attempt * attempt * 10)
    
    #Creates and fills title_metadata, descriptive_metadata, and kbart_metadata spreadsheets.
    def process_records(self):
        self.records = {}
        self.title_metadata = {}
        self.descriptive_metadata = {}
        self.kbart_metadata = {}
        self.title_metadata['header'] = {'title_name' : 'Title Name', 'title_uuid' : 'Title UUID', 'title_material_type' : 'Title Material Type', 'title_format' : 'Title Format', 'title_oclc' : 'Title OCLC', 'title_digital_holding_range' : 'Title Digital Holding Range', 'title_resolution' : 'Title Resolution', 'title_color_depth' : 'Title Color Depth', 'title_location_code' : 'Title Location Code', 'title_catalog_link' : 'Title Catalog Link', 'title_external_link' : 'Title External Link', 'item_name' : 'Item Name', 'item_uuid' : 'Item UUID', 'item_pub_year' : 'Item Pub Year', 'item_pub_month' : 'Item Pub Month', 'item_pub_day' : 'Item Pub Day', 'item_group' : 'Item Group', 'item_ocr' : 'Item OCR', 'item_coordinated_ocr' : 'Item Coordinated OCR', 'item_pdf' : 'Item PDF', 'item_access' : 'Item Access'}
        self.descriptive_metadata['header'] = {'title_uuid' : 'title uuid', 'field' : 'field', 'value' : 'value'}
        self.kbart_metadata['header'] = {'publication_title' : 'publication_title', 'print_identifier' : 'print_identifier', 'online_identifier' : 'online_identifier', 'date_first_issue_online' : 'date_first_issue_online', 'num_first_vol_online' : 'num_first_vol_online', 'num_first_issue_online' : 'num_first_issue_online', 'date_last_issue_online' : 'date_last_issue_online', 'num_last_vol_online' : 'num_last_vol_online', 'num_last_issue_online' : 'num_last_issue_online', 'title_url' : 'title_url', 'first_author' : 'first_author', 'title_id' : 'title_id', 'embargo_info' : 'embargo_info', 'coverage_depth' : 'coverage_depth', 'coverage_notes' : 'coverage_notes', 'publisher_name' : 'publisher_name', 'location' : 'location', 'title_notes' : 'title_notes', 'staff_notes' : 'staff_notes', 'vendor_id' : 'vendor_id', 'oclc_collection_name' : 'oclc_collection_name', 'oclc_collection_id' : 'oclc_collection_id', 'oclc_entry_id' : 'oclc_entry_id', 'oclc_linkscheme' : 'oclc_linkscheme', 'oclc_number' : 'oclc_number', 'action' : 'ACTION'}
        title_items_row = 0
        items_row = 0
        descriptive_metadata_row = 0
        kbart_metadata_row = 0
        for input_value in self.input_values:
            title_items = None
            descriptive_items = None
            kbart_items = None
            if self.record_source == 'worldcat':
                if input_value[2] == '':
                    marc_record, title_items, descriptive_items, kbart_items = process_oclc(input_value[0], input_value[1])
                else:
                    marc_record, title_items, descriptive_items, kbart_items = process_oclc(input_value[0], input_value[1], input_value[2])
            elif self.record_source == 'folio':
                if input_value[2] == '':
                    marc_record, title_items, descriptive_items, kbart_items = process_folio_bib_num(input_value[0], input_value[1])
                else:
                    marc_record, title_items, descriptive_items, kbart_items = process_folio_bib_num(input_value[0], input_value[1], input_value[2])
            elif self.record_source == 'folio_oclc':
                if input_value[2] == '':
                    marc_record, title_items, descriptive_items, kbart_items = get_folio_record_oclc(input_value[0], input_value[1])
                else:
                    marc_record, title_items, descriptive_items, kbart_items = get_folio_record_oclc(input_value[0], input_value[1], input_value[2])
            else:
                if input_value[2] == '':
                    marc_record, title_items, descriptive_items, kbart_items = process_millennium_bib_num(input_value[0], input_value[1])
                else:
                    marc_record, title_items, descriptive_items, kbart_items = process_millennium_bib_num(input_value[0], input_value[1], input_value[2])
            self.records[input_value[0]] = marc_record
            self.title_metadata[title_items_row] = title_items
            items_row += 1
            title_items_row += 1
            for key in descriptive_items:
                self.descriptive_metadata[descriptive_metadata_row] = descriptive_items[key]
                descriptive_metadata_row += 1
            if kbart_items is not None:
                self.kbart_metadata[kbart_metadata_row] = kbart_items
                kbart_metadata_row += 1
        self.set_up_template('title_metadata', self.title_metadata)
        self.set_up_template('descriptive_metadata', self.descriptive_metadata)
        self.set_up_template('kbart_metadata', self.kbart_metadata)
    #Opens export dialog box.
    def export(self):
        export_dialog(self)
    #Terminates process and closes window
    def teminate(self):
        root.destroy()

#Removes end punctuation and spaces.
def fix_end_char(text):
    output = text
    if re.search('(.*[^\\\.\s\;\,\:\/])([\\\.\s\;\,\:\/]+$)', text):
        output =  re.search('(.*[^\\\.\s\;\,\:\/])([\\\.\s\;\,\:\/]+$)', text).group(1)
    return output

#Removes duplicates from list while preserving order
def unique(items):
    found = set()
    keep = []
    for item in items:
        if item not in found:
            found.add(item)
            keep.append(item)
    return keep

#Removes the prefix before the numerical value
def remove_prefix(text):
    output = text
    while re.search('(?:\([Oo][Cc][Oo][Ll][Cc]\)|[Oo][Cc][Mm]|[Oo][Cc][Nn]|[Oo][Nn])', output):
        if re.search('(.*)(?:\([Oo][Cc][Oo][Ll][Cc]\)|[Oo][Cc][Mm]|[Oo][Cc][Nn]|[Oo][Nn])(.+$)', output):
            output = re.search('(.*)(?:\([Oo][Cc][Oo][Ll][Cc]\)|[Oo][Cc][Mm]|[Oo][Cc][Nn]|[Oo][Nn])(.+$)', output).group(1) + re.search('(.*)(?:\([Oo][Cc][Oo][Ll][Cc]\)|[Oo][Cc][Mm]|[Oo][Cc][Nn]|[Oo][Nn])(.+$)', output).group(2)
    return output

#Removes all subfields other than subfield a
def remove_subfields(input_field):
    out_field = pymarc.Field(tag = input_field.tag, indicators = [input_field.indicators[0], input_field.indicators[1]])
    sub_last = len(input_field.subfields)
    sub = 0
    while input_field.subfields != None and sub < sub_last:
        if re.search('(^[a]$)', input_field.subfields[sub][0]):
            inst = re.search('(^[a]$)', input_field.subfields[sub][0]).group(1)
            temp_subfield = input_field.get_subfields(inst)
            for occurrence in range(len(temp_subfield)):
                out_field.add_subfield(inst, temp_subfield[occurrence])
            for occurrence in temp_subfield:
                input_field.delete_subfield(inst)
            sub_last = len(input_field.subfields)
        else:
            sub += 1
    return out_field

#Removes given subfield
def remove_subfield(input_field, sub_id):
    out_field = pymarc.Field(tag = input_field.tag, indicators = [input_field.indicators[0], input_field.indicators[1]])
    sub_last = len(input_field.subfields)
    sub = 0
    while input_field.subfields != None and sub < sub_last:
        sub_search = re.search('(^' + sub_id+ '$)', input_field.subfields[sub][0])
        if sub_search:
            inst = sub_search.group(1)
            temp_subfield = input_field.get_subfields(inst)
            for occurrence in temp_subfield:
                input_field.delete_subfield(inst)
            sub_last = len(input_field.subfields)
        else:
            out_field.add_subfield(input_field.subfields[sub], input_field.subfields[sub + 1])
            sub += 2
    return out_field


#Modifies subfields e and 4
def format_author_field(input_field):
    out_field = pymarc.Field(tag = input_field.tag, indicators = [input_field.indicators[0], input_field.indicators[1]])
    sub_dups = []
    sub_last = len(input_field.subfields)
    sub = 0
    while input_field.subfields != None and sub < sub_last:
        if re.search('(^[e]$)', input_field.subfields[sub][0]):
            inst = re.search('(^[e]$)', input_field.subfields[sub][0]).group(1)
            temp_subfield = input_field.get_subfields(inst)
            for occurrence in range(len(temp_subfield)):
                if not re.search('(author)', temp_subfield[occurrence]):
                    fixed = fix_end_char(temp_subfield[occurrence]).lower()
                    if fixed in author_dict:
                        fixed = author_dict[fixed]
                    if fixed not in sub_dups:
                        out_field.add_subfield(inst, '(' + fixed + ')')
                        sub_dups.append(fixed)
            for occurrence in temp_subfield:
                input_field.delete_subfield(inst)
            sub_last = len(input_field.subfields)
        elif re.search('(^[4]$)', input_field.subfields[sub][0]):
            inst = re.search('(^[4]$)', input_field.subfields[sub][0]).group(1)
            temp_subfield = input_field.get_subfields(inst)
            for occurrence in range(len(temp_subfield)):
                if not re.match('(^aut)', temp_subfield[occurrence]):
                    fixed = author_dict[fix_end_char(temp_subfield[occurrence])]
                    if fixed not in sub_dups:
                        out_field.add_subfield(inst, '(' + fixed + ')')
                        sub_dups.append(fixed)
            for occurrence in temp_subfield:
                input_field.delete_subfield(inst)
            sub_last = len(input_field.subfields)
        else:
            if sub + 2 < sub_last:
                if re.search('([^A-Za-z]+[A-Za-z]\.[^A-Za-z\d]*$)', input_field.subfields[sub + 1]):
                   out_field.add_subfield(input_field.subfields[sub], fix_end_char(input_field.subfields[sub + 1]) + '.,')
                else:
                   out_field.add_subfield(input_field.subfields[sub], fix_end_char(input_field.subfields[sub + 1]) + ',')
            else:
                if re.search('([^A-Za-z]+[A-Za-z]\.[^A-Za-z\d]*$)', input_field.subfields[sub + 1]):
                   out_field.add_subfield(input_field.subfields[sub], input_field.subfields[sub + 1])
                else:
                   out_field.add_subfield(input_field.subfields[sub], fix_end_char(input_field.subfields[sub + 1]))
            sub += 2
    return out_field

def fix_245_field(input_field):
    out_field = pymarc.Field(tag = '245', indicators = [input_field.indicators[0], input_field.indicators[1]])
    sub_last = len(input_field.subfields)
    sub = 0
    while input_field.subfields != None and sub < sub_last:
        if re.search('(^[abpn]$)', input_field.subfields[sub][0]):
            if re.search('(^[a]$)', input_field.subfields[sub][0]) and re.search('(\")', input_field.subfields[sub + 1]):
                if input_field['h'] is not None:
                    for subf in input_field.get_subfields('h'):
                        if re.search('(\")', subf):
                            input_field.subfields[sub + 1] = input_field.subfields[sub + 1] + '"'
            inst = re.search('(^[abpn]$)', input_field.subfields[sub][0]).group(1)
            temp_subfield = input_field.get_subfields(inst)
            for occurrence in range(len(temp_subfield)):
                out_field.add_subfield(inst, temp_subfield[occurrence])
            for occurrence in temp_subfield:
                input_field.delete_subfield(inst)
        else:
            inst = re.search('(^.$)', input_field.subfields[sub][0]).group(1)
            temp_subfield = input_field.get_subfields(inst)
            for occurrence in temp_subfield:
                input_field.delete_subfield(inst)
        sub_last = len(input_field.subfields)
    return out_field

#Adds a leading zero if given a single digit number, returns the original number otherwise.
def pad(num):
    num_str = str(num)
    if len(num_str) == 1:
        num_str = '0' + num_str
    return num_str

#Gets the date with no separators: yyyymmdd
def get_date():
    output = str(time.localtime().tm_year) + pad(time.localtime().tm_mon) + pad(time.localtime().tm_mday)
    return output

#Gets oclc number from 035a if record source is Folio and 001 otherwise.
def get_oclc_number(record, record_source):
    oclc_num = ''
    if (record_source != 'folio' and record_source != 'folio_oclc') and record['001'] is not None:
        for f in record.get_fields('001'):
            if re.search('(?:\([Oo][Cc][Oo][Ll][Cc]\)|[Oo][Cc][Mm]|[Oo][Cc][Nn]|[Oo][Nn])', f.value()):
                oclc_num = remove_prefix(f.value())
                if re.search('(^\d+)(?:\s+$)', oclc_num):
                    oclc_num = re.search('(^\d+)(?:\s+$)', oclc_num).group(1)
                    break
                elif re.search('(^\d+$)', oclc_num):
                    break
            elif re.search('(^\d+\s*$)', f.value()):
                if re.search('(^\d+)(?:\s+$)', f.value()):
                    oclc_num = re.search('(^\d+)(?:\s+$)', f.value()).group(1)
                    break
                oclc_num = f.value()
                break
    elif record['035'] is not None and record['035']['a'] is not None:
        for f in record.get_fields('035'):
            if f['a'] is not None:
                if re.search('(?:\([Oo][Cc][Oo][Ll][Cc]\)|[Oo][Cc][Mm]|[Oo][Cc][Nn]|[Oo][Nn])', f['a']):
                    oclc_num = remove_prefix(f['a'])
                    if re.search('(^\d+)(?:\s+$)', oclc_num):
                        oclc_num = re.search('(^\d+)(?:\s+$)', oclc_num).group(1)
                        break
                    if re.search('(^\d+$)', oclc_num):
                        break
                elif re.search('(^\d+\s*$)', f['a']):
                    if re.search('(^\d+)(?:\s+$)', f['a']):
                        oclc_num = re.search('(^\d+)(?:\s+$)', f['a']).group(1)
                        break
                    oclc_num = f['a']
                    break
    return oclc_num

color_depth_dict = {'1' : '1 bit bitonal', '8' : '8 bit grayscale', '24' : '24 bit color', '(8, 8, 8)' : '24 bit color', '(8, 8, 8, 8)' : '32 bit color'}

#Gets the resolution and color_depth from tiff files.
def get_scan_data(image_file):
    tif = tifffile.TiffFile(image_file)
    if 'BitsPerSample' not in tif.pages[0].tags:
        return [tif.pages[0].tags['XResolution'].value[0] / tif.pages[0].tags['XResolution'].value[1], str(1)]
    else:
        return [tif.pages[0].tags['XResolution'].value[0] / tif.pages[0].tags['XResolution'].value[1], str(tif.pages[0].tags['BitsPerSample'].value)]


#Extract marc record for given oclc number from Worldcat.
def process_oclc(oclc_num, input_uuid, input_collection = None, attempt = 1):
    html_url = 'http://www.worldcat.org/webservices/catalog/content/' + oclc_num + '?servicelevel=full&wskey=' + api_keys.get_apikey()
    req = urllib.request.Request(html_url, headers={'User-Agent': 'Mozilla/5.0'})
    response = urllib.request.urlopen(req, timeout=3.0)
    if response.status == 200:
        marc_record_worldcat = response.read()
        marc_record = next(text_marc_reader.get_marc_worldcat_xml(marc_record_worldcat))
        return process_marc_file(marc_record, None, input_uuid, input_collection, record_source='worldcat')
    else:
        if attempt <= 10:
            time.sleep(2 * attempt)
            return process_oclc(oclc_num, input_uuid, input_collection, attempt = attempt + 1)
        else:
            print('Failed to get record for oclc number ' + str(oclc_num) + ' after ' + str(attempt) + 'attempts.')
            return None

#Extract marc record for given bib number from Millennium.
def process_millennium_bib_num(input_bib_num, input_uuid, input_collection = None):
    bib_num = ''
    if len(str(input_bib_num)) == 9:
        bib_num = input_bib_num[:8]
    elif len(str(input_bib_num)) == 8:
        bib_num = input_bib_num
    html_url = 'http://catalog-old.crl.edu/search~S1?/.' + bib_num + '/.' + bib_num + '/1%2C1%2C1%2CB/marc~' + bib_num
    req = urllib.request.Request(html_url, headers={'User-Agent': 'Mozilla/5.0'})
    webpage = urllib.request.urlopen(req).read()
    soup = BeautifulSoup(webpage, 'html.parser', from_encoding='utf-8')
    marc = soup.find_all('pre')
    marc_record_millennium = ''
    for a in marc:
        if re.match('\nLEADER', a.getText()):
            marc_record_millennium = a.getText()
            break
    marc_record = next(text_marc_reader.get_marc_millennium((marc_record_millennium)))
    return process_marc_file(marc_record, input_bib_num, input_uuid, input_collection, record_source='millennium')

#Extract marc record for given uuid number from Folio.
def process_folio_bib_num(folio_uuid, input_uuid, input_collection = None):
    marc_record = get_marc(folio_uuid)
    return process_marc_file(marc_record, folio_uuid, input_uuid, input_collection, record_source='folio')


#Extract marc record for given oclc number from Folio.
def get_folio_record_oclc(oclc_num, input_uuid, input_collection = None):
    marc_record = get_marc_record_from_oclc(oclc_num)
    return process_marc_file(marc_record, oclc_num, input_uuid, input_collection, record_source='folio_oclc')


#Returns the record material type and collection.
def get_material_type_folio(record, oclc_number):
    record_type = record.leader[7]
    material_type = None
    collection = None
    if re.search('([mi])', record_type):
        collection = 'monograph'
        dissertation = True
        material_type = 1
        html_url = 'https://catalog.crl.edu/Search/Results?lookfor=' + oclc_number + '&type=oclc_num&filter%5B%5D=crl_scope%3A"Dissertations"'
        req = urllib.request.Request(html_url, headers={'User-Agent': 'Mozilla/5.0'})
        webpage = urllib.request.urlopen(req).read()
        soup = BeautifulSoup(webpage, 'html.parser', from_encoding='utf-8')
        web_contents = soup.find_all('div')
        for div in web_contents:
            if 'class' in list(div.attrs) and div['class'] == ['mainbody', 'left']:
                div_seach = div.find_all('h2')
                for h2 in div_seach:
                    if h2.get_text() == 'No Results!':
                        dissertation = False
                        break
                if not dissertation:
                    break
        if dissertation:
            material_type = 6
    elif re.search('(s)', record_type):
        collection = 'serial'
        serial = True
        html_url = 'https://catalog.crl.edu/Search/Results?lookfor=' + oclc_number + '&type=oclc_num&filter%5B%5D=crl_scope%3A"Serials"'
        req = urllib.request.Request(html_url, headers={'User-Agent': 'Mozilla/5.0'})
        webpage = urllib.request.urlopen(req).read()
        soup = BeautifulSoup(webpage, 'html.parser', from_encoding='utf-8')
        web_contents = soup.find_all('div')
        for div in web_contents:
            if 'class' in list(div.attrs) and div['class'] == ['mainbody', 'left']:
                div_seach = div.find_all('h2')
                for h2 in div_seach:
                    if h2.get_text() == 'No Results!':
                        serial = False
                        break
                if not serial:
                    break
        if serial:
            material_type = 2
        else:
            collection = 'newspaper'
            newspaper = True
            html_url = 'https://catalog.crl.edu/Search/Results?lookfor=' + oclc_number + '&type=oclc_num&filter%5B%5D=crl_scope%3A"Newspapers"'
            req = urllib.request.Request(html_url, headers={'User-Agent': 'Mozilla/5.0'})
            webpage = urllib.request.urlopen(req).read()
            soup = BeautifulSoup(webpage, 'html.parser', from_encoding='utf-8')
            web_contents = soup.find_all('div')
            for div in web_contents:
                if 'class' in list(div.attrs) and div['class'] == ['mainbody', 'left']:
                    div_seach = div.find_all('h2')
                    for h2 in div_seach:
                        if h2.get_text() == 'No Results!':
                            newspaper = False
                            break
                    if not newspaper:
                        break
            if newspaper:
                material_type = 3
    elif re.search('(c)', record_type):
        collection = 'monograph'
        material_type = 4
    government_publication = record['008'].value()[28]
    if not re.match('(u$|\s$|\\|$)', government_publication):
        material_type = 5
    return [material_type, collection]

#Removes ending comma from string.
def remove_ending_comma(text):
    if re.match('(.+)(,\s*$)', text):
        text = re.match('(.+)(,\s*$)', text).group(1)
    return text

#Extracts title_items, descriptive_items, kbart_items from marc_record.
#Returns marc_record, title_items, descriptive_items, kbart_items.
def process_marc_file(marc_record, input_bib_num, input_uuid, input_collection = None, record_source='millennium', extention='.csv'):
    record = marc_record
    catalog_link = None
    bib_num_id = None
    print(record)
    electronic =  record['008'] is not None and record['008'].data[23] == 'o'
    if input_bib_num is not None:
        bib_num_id = input_bib_num
        if (record_source == 'folio' or record_source == 'folio_oclc') and electronic:
            catalog_link = 'https://catalog.crl.edu/Record/' + input_bib_num
    if record_source == 'folio_oclc' and record['999'] is not None and record['999']['i'] is not None:
        bib_num_id = record['999']['i']
    oclc_number = ''
    country = False
    languages = []
    subjects = []
    publisher = []
    row = 0
    ddsnext_uuid = input_uuid
    title_items = {'title_name' : None, 'title_uuid' : ddsnext_uuid, 'title_material_type' : None, 'title_format' : None, 'title_oclc' : None, 'title_digital_holding_range' : None, 'title_resolution' : None, 'title_color_depth' : None, 'title_location_code' : None, 'title_catalog_link' : catalog_link, 'title_external_link' : None, 'item_name' : None, 'item_uuid' : None, 'item_pub_year' : None, 'item_pub_month' : None, 'item_pub_day' : None, 'item_group' : None, 'item_ocr' : None, 'item_coordinated_ocr' : None, 'item_pdf' : None, 'item_access' : None}
    descriptive_items = {}
    kbart_items = {'publication_title' : None, 'print_identifier' : None, 'online_identifier' : None, 'date_first_issue_online' : None, 'num_first_vol_online' : None, 'num_first_issue_online' : None, 'date_last_issue_online' : None, 'num_last_vol_online' : None, 'num_last_issue_online' : None, 'title_url' : None, 'first_author' : None, 'title_id' : None, 'embargo_info' : None, 'coverage_depth' : None, 'coverage_notes' : None, 'publisher_name' : None, 'location' : None, 'title_notes' : None, 'staff_notes' : None, 'vendor_id' : None, 'oclc_collection_name' : None, 'oclc_collection_id' : None, 'oclc_entry_id' : bib_num_id, 'oclc_linkscheme' : None, 'oclc_number' : None, 'action' : 'raw'}
    #If record is monograph, extracts item_pub_year, item_ocr, item_coordinated_ocr, item_pdf, item_pub_year, item_access, item_name
    if input_collection is not None and not electronic:
        input_collection = remove_returns(input_collection)
        kbart_items['coverage_depth'] = collection_dict[input_collection]['coverage_depth']
        kbart_items['oclc_collection_name'] = collection_dict[input_collection]['oclc_collection_name']
        kbart_items['oclc_collection_id'] = collection_dict[input_collection]['oclc_collection_id']
        if re.match('(monograph$)', input_collection.lower()):
            pub_year = None
            if record['260'] is not None and record['260']['c'] is not None:
                pub_year = record['260']['c']
            elif record['264'] is not None and record['264']['c'] is not None:
                pub_year = record['264']['c']
            if pub_year is not None and re.search('(\d{4})', pub_year):
                kbart_items['date_first_issue_online'] = re.search('(\d{4})', pub_year).group(1)
                if input_collection is not None and input_collection == 'monograph':
                    title_items['item_pub_year'] = re.search('(\d{4})', pub_year).group(1)
                    title_items['item_ocr'] = 1
                    title_items['item_coordinated_ocr'] = 1
                    title_items['item_pdf'] = 1
                    if int(title_items['item_pub_year']) <= 1926:
                        title_items['item_access'] = 1
                    else:
                        title_items['item_access'] = 2
        if input_collection == 'monograph':
            title_items['item_name'] = 'Full text'
    #Extracts title
    if record['245'] is not None:
        output_title = fix_end_char(fix_245_field(record['245']).value())
        title_items['title_name'] = output_title
        kbart_items['publication_title'] = output_title
    #Extracts title
    elif record['222'] is not None:
        output_title = remove_subfield(record['222'], 'b').value()
        title_items['title_name'] = output_title
        kbart_items['publication_title'] = output_title
    #Extracts and prints series title (CRL collection title variant)
    if record['246'] is not None:
        if record['246']['i'] is not None and record['246']['a'] is not None and re.search('CRL collection title', record['246']['i']):
            descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'series', 'value' : record['246']['i'] + ' ' + record['246']['a']}
            row += 1
    #Extracts title_location_code
    if record['998'] is not None and record['998']['a'] is not None:
        location = ''
        for sub in record['998'].get_subfields('a'):
            if re.match('(camp|samp|memp|lamp|seam|seem)([^e])', sub):
                location = sub
                break
            elif re.match('(crl[^e]|disse|txbke)', sub) and not re.match('(camp|samp|memp|lamp|seam|seem)', location):
                location = sub
            elif re.match('(diss)', sub) and not re.match('(camp|samp|memp|lamp|seam|seem)', location):
                location = 'disse'
            elif re.match('(txbk)', sub) and not re.match('(camp|samp|memp|lamp|seam|seem)', location):
                location = 'txbke'
            elif re.match('(fdoc[^e]|fogse)', sub) and not re.match('(camp|samp|memp|lamp|seam|seem|crl|diss|txbk)', location):
                location = sub
            elif re.match('(fogs)', sub) and not re.match('(camp|samp|memp|lamp|seam|seem|crl|diss|txbk)', location):
                location = 'fogse'
            elif re.match('(wna|grci)', sub) and not re.match('(camp|samp|memp|lamp|seam|seem|crl|diss|txbk|fdoc|fogs)', location):
                location = sub
        title_items['title_location_code'] = location
    #Extracts holdings and title url.
    if record['856'] is not None and record['856']['u'] is not None:
        for f in record.get_fields('856'):
            for sub in f.get_subfields('u'):
                if f['z'] is None or f['z'] is not None and not re.search('(?:[Gg][Uu][Ii][Dd][Ee])', f['z']):
                    if re.search('(?:.*ddsnext\.crl\.edu\/titles\/)(\d+)', sub):
                        kbart_items['title_url'] = sub
                        if f['z'] is not None and re.search('(?:.\:\s*)(.+)', f['z']):
                            title_items['title_digital_holding_range'] = re.search('(?:.\:\s*)(.+)', f['z']).group(1)
                        elif f['3'] is not None and re.search('(?:.\:\s*)(.+)', f['3']):
                            title_items['title_digital_holding_range'] = re.search('(?:.\:\s*)(.+)', f['3']).group(1)
    #Extracts and prints LCCN
    if record['010'] is not None and record['010']['a'] is not None and electronic:
        descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'LCCN', 'value' : record['010']['a']}
        row += 1
    #Extracts and prints ISBN
    if record['020'] is not None and electronic:
        descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'ISBN', 'value' : record['020'].value()}
        kbart_items['online_identifier'] = record['020']
        row += 1
    #Extracts and prints ISSN
    if record['022'] is not None and record['022']['a'] is not None and electronic:
        descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'ISSN', 'value' : record['022']['a']}
        kbart_items['online_identifier'] = record['022']['a']
        row += 1
    #Extracts and prints ISSN
    if record['022'] is not None and record['022']['l'] is not None and electronic:
        kbart_items['print_identifier'] = record['022']['l']
    #Extracts and prints OCLC number
    oclc_number = remove_whitespace(get_oclc_number(record, record_source))
    if oclc_number != '' and electronic:
        title_items['title_oclc'] = oclc_number
        kbart_items['oclc_number'] = oclc_number
    #Extracts material type
    if record.leader is not None:
        if record_source == 'millennium':
            title_items['title_material_type'], input_collection = get_material_type_millennium(record, oclc_number)
        elif record_source == 'folio' or record_source == 'folio_oclc':
            title_items['title_material_type'], input_collection = get_material_type_folio(record, oclc_number)
        if title_items['title_material_type'] is not None:
            kbart_items['coverage_depth'] = collection_dict[input_collection]['coverage_depth']
            kbart_items['oclc_collection_name'] = collection_dict[input_collection]['oclc_collection_name']
            kbart_items['oclc_collection_id'] = collection_dict[input_collection]['oclc_collection_id']
            if re.match('(monograph$)', input_collection.lower()) and record['008'] is not None and record['008'].data[23] != 'o':
                pub_year = None
                if record['260'] is not None and record['260']['c'] is not None:
                    pub_year = record['260']['c']
                elif record['264'] is not None and record['264']['c'] is not None:
                    pub_year = record['264']['c']
                if pub_year is not None and re.search('(\d{4})', pub_year):
                    kbart_items['date_first_issue_online'] = re.search('(\d{4})', pub_year).group(1)
                    if input_collection is not None and input_collection == 'monograph':
                        title_items['item_pub_year'] = re.search('(\d{4})', pub_year).group(1)
                        title_items['item_ocr'] = 1
                        title_items['item_coordinated_ocr'] = 1
                        title_items['item_pdf'] = 1
                        if int(title_items['item_pub_year']) <= 1926:
                            title_items['item_access'] = 1
                        else:
                            title_items['item_access'] = 2
                title_items['item_name'] = 'Full text'
    #Extracts and prints dissertation description
    if record['502'] is not None:
        description = ''
        #Extracts the dissertation description from the 502$a
        if record['502']['a'] is not None:
            description = record['502']['a']
            if re.match('(?:.*\()(.+)(?:\)\-\-)', record['502']['a']):
                description = description + re.match('(?:.*\()(.+)(?:\)\-\-)', record['502']['a']).group(1)
            if re.match('(?:.*\([^\)]+\)\-\-)(.+)(?:\,)(\d{4})(?:\S*$)', record['502']['a']):
                if description != '':
                    description = description + ' '
                description = description + re.match('(?:.*\([^\)]+\)\-\-)(.+)(?:\,)(\d{4})(?:\S*$)', record['502']['a']).group(1) + ' ' + re.match('(?:.*\([^\)]+\)\-\-)(.+)(?:\,)(\d{4})(?:\S*$)', record['502']['a']).group(2)
            elif re.match('(?:.*\([^\)]+\)\-\-)(.+)(?:\S*$)', record['502']['a']):
                if description != '':
                    description = description + ' '
                description = description + re.match('(?:.*\([^\)]+\)\-\-)(.+)(?:\S*$)', record['502']['a']).group(1)
        #Extracts the dissertation description from the 502$b, 502$c, 502$d, and 502$g
        else:
            if record['502']['g'] is not None:
                description = record['502']['g']
            if record['502']['b'] is not None:
                if description != '':
                    description = description + ' '
                description = description + record['502']['b']
            if record['502']['c'] is not None:
                if description != '':
                    description = description + ' '
                description = description + record['502']['c']
            if record['502']['d'] is not None:
                if description != '':
                    description = description + ' '
                description = description + record['502']['d']
        descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'description', 'value' : description}
        row += 1
    #Extracts and prints description
    if record['520'] is not None and record['520']['a'] is not None:
        description = record['520']['a']
        if record['520']['b'] is not None:
            if description != '':
                description = description + ' '
            description = description + record['520']['b']
        descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'description', 'value' : description}
        row += 1
    #Extracts language
    if record['041'] is not None:
        if record['041']['a'] is not None:
            for sub in record['041'].get_subfields('a'):
                if sub in language_dict:
                    languages.append(language_dict[sub])
    #Extracts country
    #Extracts language
    if record['998'] is not None:
        if record['998']['g'] is not None:
            if fix_end_char(record['998']['g']) in country_dict and fix_end_char(record['998']['g']) != 'xx':
                descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'country', 'value' : country_dict[fix_end_char(record['998']['g'])]}
                row += 1
                country = True
        if record['998']['f'] is not None:
            if record['998']['f'] in language_dict:
                if record['041'] is None or record['041'] is not None and (record['041']['a'] is None or (record['041']['a'] is not None and record['998']['f'] not in record['041'].get_subfields('a'))):
                    languages.append(language_dict[record['998']['f']])
    #Extracts and prints country if not found in the 998 field
    #Extracts language
    if record['008'] is not None:
        if not country:
            if re.search('([a-zA-Z]+)', record['008'].value()[15:18]):
                country_code = re.search('([a-zA-Z]+)', record['008'].value()[15:18]).group(1)
            if fix_end_char(country_code) in country_dict and fix_end_char(country_code) != 'xx':
                descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'country', 'value' : country_dict[fix_end_char(country_code)]}
                row += 1
        if record['008'].value()[35:38] in language_dict:
            languages.append(language_dict[record['008'].value()[35:38]])
    #Deduplicate languages
    if languages != []:
        languages = unique(languages)
    #Prints languages
    for language in languages:
        descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'language', 'value' : language}
        row += 1
    #Extracts and prints coverage (country)
    if record['752'] is not None and record['752']['a'] is not None:
        descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'coverage', 'value' : record['752']['a']}
        row += 1
    #Extracts and prints call number if record is an electronic Millennium record.
    if record['099'] is not None:
        if record['099']['a'] is not None:
            for sub in record['099'].get_subfields('a'):
                if sub != 'Internet resource' and sub != 'ediss' and sub != 'MF' and sub != 'Electronic version' and sub != 'Electronic resource/e' and sub != 'TOSS' and not re.search('[Ee][Ll][Ee][Cc]?[Tt][Rr][Oo][Nn][Ii][Cc]', sub):
                    if electronic:
                        descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'resource_identifier', 'value' : sub}
                    if input_collection is not None and input_collection == 'monograph' and record['008'] is not None and record['008'].data[23] != 'o':
                        item_uuid = sub
                        if re.match('(.+)(?:\s+$)', item_uuid):
                            item_uuid = re.match('(.+)(?:\s+$)', item_uuid).group(1)
                        title_items['item_uuid'] = item_uuid.replace('/', '-')
                    row += 1
    #Extracts and prints creator
    #Extracts subject (title)
    if record['100'] is not None:
        for f in record.get_fields('100'):
            if f['t'] is not None:
                subjects.append(f['t'])
                f.delete_subfield('t')
            author = remove_ending_comma(remove_subfield(format_author_field(f), '6').value())
            descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'creator', 'value' : author}
            row += 1
            kbart_items['first_author'] = remove_ending_comma(record['100']['a'])
    #Extracts and prints creator
    #Extracts subject (title)
    if record['110'] is not None:
        for f in record.get_fields('110'):
            if f['t'] is not None:
                subjects.append(f['t'])
                f.delete_subfield('t')
            author = remove_subfield(format_author_field(f), '6').value()
            descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'creator', 'value' : author}
            row += 1
    #Extracts and prints creator
    #Extracts subject (title)
    if record['111'] is not None:
        for f in record.get_fields('111'):
            if f['t'] is not None:
                subjects.append(f['t'])
                f.delete_subfield('t')
            author = remove_subfield(format_author_field(f), '6').value()
            descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'creator', 'value' : author}
            row += 1
    #Extracts and prints creator
    #Extracts subject (title)
    if record['700'] is not None:
        for f in record.get_fields('700'):
            if f['t'] is not None:
                subjects.append(f['t'])
                f.delete_subfield('t')
            author = remove_ending_comma(remove_subfield(format_author_field(f), '6').value())
            descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'creator', 'value' : author}
            row += 1
    #Extracts and prints creator
    #Extracts subject (title)
    if record['710'] is not None:
        for f in record.get_fields('710'):
            if f['t'] is not None:
                subjects.append(f['t'])
                f.delete_subfield('t')
            author = remove_subfield(format_author_field(f), '6').value()
            descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'creator', 'value' : author}
            row += 1
    #Extracts and prints creator
    #Extracts subject (title)
    if record['711'] is not None:
        for f in record.get_fields('711'):
            if f['t'] is not None:
                subjects.append(f['t'])
                f.delete_subfield('t')
            author = remove_subfield(format_author_field(f), '6').value()
            descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'creator', 'value' : author}
            row += 1
    #Extracts publishers
    if record['260'] is not None:
        for f in record.get_fields('260'):
            if f['b'] is not None:
                for sub in f.get_subfields('b'):
                    if not (re.search('(\[?publisher not identified\]?)', sub) or re.search('(\[?[Ss]\.n\.\??\]?)', sub)):
                        publisher.append(fix_end_char(sub))
    #Extracts publishers
    elif record['264'] is not None:
        for f in record.get_fields('264'):
            if f.indicator2 == '1' and f['b'] is not None:
                for sub in f.get_subfields('b'):
                    if not (re.search('(\[?publisher not identified\]?)', sub) or re.search('(\[?s\.n\.\]?)', sub)):
                        publisher.append(fix_end_char(sub))
    #Deduplicate publishers
    if publisher != []:
        publisher = unique(publisher)
    #Prints publishers
    for pub in publisher:
        descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'publisher', 'value' : pub}
        row += 1
        if kbart_items['publisher_name'] is None:
            kbart_items['publisher_name'] = pub
    #Extracts and prints series
    #Extracts and prints author
    if record['800'] is not None:
        for f in record.get_fields('800'):
            f = remove_subfield(f, '6')
            if f['t'] is not None:
                descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'series', 'value' : f['a']}
                row += 1
                f.delete_subfield('t')
            if f['v'] is not None:
                f.delete_subfield('v')
            author = remove_ending_comma(remove_subfield(format_author_field(f), '6').value())
            descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'creator', 'value' : author}
            row += 1
    #Extracts and prints series
    #Extracts and prints author
    if record['810'] is not None:
        for f in record.get_fields('810'):
            f = remove_subfield(f, '6')
            if f['t'] is not None:
                descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'series', 'value' : f['a']}
                row += 1
                f.delete_subfield('t')
            if f['v'] is not None:
                f.delete_subfield('v')
            author = remove_subfield(format_author_field(f), '6').value()
            descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'creator', 'value' : author}
            row += 1
    #Extracts and prints series
    #Extracts and prints author
    if record['811'] is not None:
        for f in record.get_fields('811'):
            f = remove_subfield(f, '6')
            if f['t'] is not None:
                descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'series', 'value' : f['a']}
                row += 1
                f.delete_subfield('t')
            if f['v'] is not None:
                f.delete_subfield('v')
            author = remove_subfield(format_author_field(f), '6').value()
            descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'creator', 'value' : author}
            row += 1
    #Extracts and prints series
    if record['830'] is not None:
        for f in record.get_fields('830'):
            f = remove_subfield(f, '6')
            if f['a'] is not None:
                descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'series', 'value' : f['a']}
                row += 1
    #Extracts and prints subject (author)
    #Extracts and prints subject (title)
    if record['600'] is not None:
        for f in record.get_fields('600'):
            f = remove_subfield(f, '6')
            if f.indicator2 == '0':
                if f['t'] is not None:
                    subjects.append(f['t'])
                    f.delete_subfield('t')
                if f['v'] is not None:
                    subjects.append(f['v'])
                    f.delete_subfield('v')
                subjects.append(format_author_field(f).value())
    #Extracts and prints subject (author)
    #Extracts and prints subject (title)
    if record['610'] is not None:
        for f in record.get_fields('610'):
            f = remove_subfield(f, '6')
            if f.indicator2 == '0':
                if f['t'] is not None:
                    subjects.append(f['t'])
                    f.delete_subfield('t')
                if f['v'] is not None:
                    subjects.append(f['v'])
                    f.delete_subfield('v')
                subjects.append(format_author_field(f).value())
    #Extracts and prints subject (author)
    #Extracts and prints subject (title)
    if record['611'] is not None:
        for f in record.get_fields('611'):
            f = remove_subfield(f, '6')
            if f.indicator2 == '0':
                if f['t'] is not None:
                    subjects.append(f['t'])
                    f.delete_subfield('t')
                if f['v'] is not None:
                    subjects.append(f['v'])
                    f.delete_subfield('v')
                subjects.append(format_author_field(f).value())
    #Extracts and prints subject
    if record['630'] is not None:
        for f in record.get_fields('630'):
            f = remove_subfield(f, '6')
            if f.indicator2 == '0':
                for sub in range(len(f.subfields)):
                    if sub % 2 != 0:
                        subjects.append(f.subfields[sub])
    #Extracts and prints subject
    if record['650'] is not None:
        for f in record.get_fields('650'):
            f = remove_subfield(f, '6')
            if f.indicator2 == '0':
                for sub in range(len(f.subfields)):
                    if sub % 2 != 0:
                        subjects.append(f.subfields[sub])
    #Extracts and prints subject
    if record['651'] is not None:
        for f in record.get_fields('651'):
            f = remove_subfield(f, '6')
            if f.indicator2 == '0':
                for sub in range(len(f.subfields)):
                    if sub % 2 != 0:
                        subjects.append(f.subfields[sub])
    #Extracts and prints original author and original title
    #Extracts subject (title)
    if record['880'] is not None:
        for f in record.get_fields('880'):
            if f['6'] is not None:
                for sub in f.get_subfields('6'):
                    #Extracts and prints original author
                    if re.search('(^[17][01][01]-)', sub):
                        if f['t'] is not None:
                            subjects.append(f['t'])
                            f.delete_subfield('t')
                        if f['v'] is not None:
                            f.delete_subfield('v')
                        orig_author = remove_subfield(format_author_field(f), '6').value()
                        if re.match('(.+)(,\s*$)', orig_author):
                            orig_author = re.match('(.+)(,\s*$)', orig_author).group(1)
                        descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'orig_author', 'value' : orig_author}
                        row += 1
                    #Extracts and prints original title
                    if re.search('(^245-)', sub):
                        orig_title = fix_end_char(fix_245_field(f).value())
                        descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'orig_title', 'value' : orig_title}
                        row += 1
    for n in range(len(subjects)):
        subjects[n] = fix_end_char(subjects[n])
    #Deduplicate subjects
    if subjects != []:
        subjects = unique(subjects)
    #Prints subjects
    for s in subjects:
        if not ((title_items['title_material_type'] == '2' or (title_items['title_material_type'] == '5' and (title_items['title_location_code'] == 'fdocs' or title_items['title_location_code'] == 'fogse'))) and fix_end_char(s) == 'Periodicals') and not (title_items['title_material_type'] == '3' and fix_end_char(s) == 'Newspapers'):
            descriptive_items[row] = {'title_uuid' : ddsnext_uuid, 'field' : 'subject', 'value' : fix_end_char(s)}
            row += 1
    if not electronic:
        kbart_items = None
    return [marc_record, title_items, descriptive_items, kbart_items]

if __name__ == "__main__":
    root = tkinter.Tk()
    app = Application(master=root)
    app.mainloop()
